import sys
import os
import uuid
import io
import openpyxl
from datetime import date, datetime, timedelta
from sqlalchemy import select

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.database import SessionLocal
from app.models.company import Company
from app.models.engineer import Engineer
from app.models.schedule import Schedule
from app.models.user import User
from app.models.performance import Performance
from fastapi.testclient import TestClient
from app.main import app
from app.services.security import create_access_token

def create_test_excel_bytes(sheet_name="Performance", headers=None, rows=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if headers:
        ws.append(headers)
    if rows:
        for r in rows:
            ws.append(r)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

def run_performance_v2_tests():
    db = SessionLocal()
    print("\n=== STARTING COMPREHENSIVE PERFORMANCE V2 TEST SUITE ===")
    
    unique_suffix = uuid.uuid4().hex[:8]

    try:
        # 1. Setup Tenant Companies A and B
        comp_a = Company(
            company_id=uuid.uuid4(),
            company_name=f"Perf V2 Corp A {unique_suffix}",
            short_name=f"PVA_{unique_suffix}",
            is_active=True
        )
        comp_b = Company(
            company_id=uuid.uuid4(),
            company_name=f"Perf V2 Corp B {unique_suffix}",
            short_name=f"PVB_{unique_suffix}",
            is_active=True
        )
        db.add_all([comp_a, comp_b])
        db.commit()

        # 2. Setup Users
        user_a = User(
            user_id=uuid.uuid4(),
            company_id=comp_a.company_id,
            email=f"admin_a_{unique_suffix}@perfv2.com",
            full_name="Perf V2 Admin A",
            role="Manager",
            password_hash="hashedpassword123",
            is_active=True
        )
        user_b = User(
            user_id=uuid.uuid4(),
            company_id=comp_b.company_id,
            email=f"admin_b_{unique_suffix}@perfv2.com",
            full_name="Perf V2 Admin B",
            role="Manager",
            password_hash="hashedpassword123",
            is_active=True
        )
        db.add_all([user_a, user_b])
        db.commit()

        # 3. Setup Engineers
        orbit_id_a = f"ORB-V2-A-{unique_suffix}"
        orbit_id_b = f"ORB-V2-B-{unique_suffix}"

        eng_a = Engineer(
            engineer_id=uuid.uuid4(),
            company_id=comp_a.company_id,
            engineer_name="V2 Engineer A",
            orbit_id=orbit_id_a,
            level="L3 Senior",
            status="Active"
        )
        eng_b = Engineer(
            engineer_id=uuid.uuid4(),
            company_id=comp_b.company_id,
            engineer_name="V2 Engineer B",
            orbit_id=orbit_id_b,
            level="L2 Specialist",
            status="Active"
        )
        db.add_all([eng_a, eng_b])
        db.commit()

        # 4. Setup Schedules
        sch_a1 = Schedule(
            schedule_id=uuid.uuid4(),
            engineer_id=eng_a.engineer_id,
            support_type="Customer Support",
            country="United States",
            fab_city="Austin",
            fab_site="Fab 1",
            start_date=date(2026, 3, 1),
            end_date=date(2026, 3, 15),
            schedule_status="Active"
        )
        sch_a2 = Schedule(
            schedule_id=uuid.uuid4(),
            engineer_id=eng_a.engineer_id,
            support_type="PM Support",
            country="Taiwan",
            fab_city="Hsinchu",
            fab_site="Fab 12",
            start_date=date(2026, 4, 1),
            end_date=date(2026, 4, 20),
            schedule_status="Upcoming"
        )
        sch_b = Schedule(
            schedule_id=uuid.uuid4(),
            engineer_id=eng_b.engineer_id,
            support_type="Tool Relocation",
            country="Germany",
            fab_city="Dresden",
            fab_site="Fab Dresden",
            start_date=date(2026, 3, 5),
            end_date=date(2026, 3, 25),
            schedule_status="Active"
        )
        db.add_all([sch_a1, sch_a2, sch_b])
        db.commit()

        # TestClient and Auth Headers
        client = TestClient(app)
        token_a = create_access_token({"sub": str(user_a.user_id)})
        headers_a = {
            "Authorization": f"Bearer {token_a}",
            "X-Company-ID": str(comp_a.company_id)
        }
        token_b = create_access_token({"sub": str(user_b.user_id)})
        headers_b = {
            "Authorization": f"Bearer {token_b}",
            "X-Company-ID": str(comp_b.company_id)
        }

        # -------------------------------------------------------------
        # SCENARIO 1: Performance Upload with Valid Schedule ID
        # -------------------------------------------------------------
        excel_valid = create_test_excel_bytes(
            sheet_name="Performance",
            headers=["Schedule ID", "Orbit ID", "Actual Start Date", "Actual End Date", "Score", "Escalation", "Escalation Reason", "Feedback"],
            rows=[
                [str(sch_a1.schedule_id), orbit_id_a, "2026-03-01", "2026-03-15", 4.8, "No", "", "Great performance on Fab 1"],
                [str(sch_a2.schedule_id), orbit_id_a, "2026-04-01", "2026-04-20", 4.2, "Yes", "Chamber gas leak incident", "Resolved chamber leak promptly"]
            ]
        )
        res1 = client.post(
            "/api/upload",
            data={"module_id": "up-performance"},
            files={"file": ("perf_valid.xlsx", excel_valid, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=headers_a
        )
        assert res1.status_code == 200, f"Upload 1 failed: {res1.text}"
        data1 = res1.json()
        assert data1["rowsProcessed"] == 2
        assert data1["errorsCount"] == 0
        print("[PASS] Scenario 1: Performance upload with valid Schedule ID succeeded.")

        # -------------------------------------------------------------
        # SCENARIO 2: Performance Upload with Nonexistent Schedule ID
        # -------------------------------------------------------------
        fake_sch_id = str(uuid.uuid4())
        excel_nonexistent = create_test_excel_bytes(
            sheet_name="Performance",
            headers=["Schedule ID", "Orbit ID", "Actual Start Date", "Score"],
            rows=[
                [fake_sch_id, orbit_id_a, "2026-05-01", 4.5]
            ]
        )
        res2 = client.post(
            "/api/upload",
            data={"module_id": "up-performance"},
            files={"file": ("perf_fake_sch.xlsx", excel_nonexistent, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=headers_a
        )
        assert res2.status_code == 200
        data2 = res2.json()
        assert data2["errorsCount"] == 1
        rep_resp2 = client.get(data2["report_url"], headers=headers_a)
        assert rep_resp2.status_code == 200
        # Verify validation report contains the clear error message
        wb2 = openpyxl.load_workbook(io.BytesIO(rep_resp2.content))
        ws_err2 = wb2["Errors"]
        err_text2 = [cell.value for cell in ws_err2['F'] if cell.value and cell.value != "Error"]
        assert any("was not found" in str(e) for e in err_text2), f"Expected 'was not found' error, got: {err_text2}"
        print("[PASS] Scenario 2: Nonexistent Schedule ID row rejected with clear error.")

        # -------------------------------------------------------------
        # SCENARIO 3: Performance Upload with Orbit ID Mismatch
        # -------------------------------------------------------------
        excel_mismatch = create_test_excel_bytes(
            sheet_name="Performance",
            headers=["Schedule ID", "Orbit ID", "Actual Start Date", "Score"],
            rows=[
                [str(sch_a1.schedule_id), orbit_id_b, "2026-03-01", 4.0]
            ]
        )
        res3 = client.post(
            "/api/upload",
            data={"module_id": "up-performance"},
            files={"file": ("perf_mismatch.xlsx", excel_mismatch, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=headers_a
        )
        assert res3.status_code == 200
        data3 = res3.json()
        assert data3["errorsCount"] == 1
        rep_resp3 = client.get(data3["report_url"], headers=headers_a)
        wb3 = openpyxl.load_workbook(io.BytesIO(rep_resp3.content))
        ws_err3 = wb3["Errors"]
        err_text3 = [cell.value for cell in ws_err3['F'] if cell.value and cell.value != "Error"]
        assert any("does not match Schedule" in str(e) for e in err_text3), f"Expected Orbit ID mismatch error, got: {err_text3}"
        print("[PASS] Scenario 3: Orbit ID mismatch rejected with clear error.")

        # -------------------------------------------------------------
        # SCENARIO 4: Performance Upload with Cross-Company Schedule
        # -------------------------------------------------------------
        excel_cross = create_test_excel_bytes(
            sheet_name="Performance",
            headers=["Schedule ID", "Orbit ID", "Actual Start Date", "Score"],
            rows=[
                [str(sch_b.schedule_id), orbit_id_b, "2026-03-05", 4.0]
            ]
        )
        res4 = client.post(
            "/api/upload",
            data={"module_id": "up-performance"},
            files={"file": ("perf_cross.xlsx", excel_cross, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=headers_a
        )
        assert res4.status_code == 200
        data4 = res4.json()
        assert data4["errorsCount"] == 1
        print("[PASS] Scenario 4: Cross-company Schedule ID upload blocked.")

        # -------------------------------------------------------------
        # SCENARIO 5: Performance Upsert using (schedule_id + actual_start_date)
        # -------------------------------------------------------------
        excel_upsert = create_test_excel_bytes(
            sheet_name="Performance",
            headers=["Schedule ID", "Orbit ID", "Actual Start Date", "Actual End Date", "Score", "Feedback"],
            rows=[
                [str(sch_a1.schedule_id), orbit_id_a, "2026-03-01", "2026-03-15", 5.0, "Updated to 5.0 rating score"]
            ]
        )
        res5 = client.post(
            "/api/upload",
            data={"module_id": "up-performance"},
            files={"file": ("perf_upsert.xlsx", excel_upsert, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=headers_a
        )
        assert res5.status_code == 200
        data5 = res5.json()
        assert data5["existingCount"] == 1
        assert data5["rowsProcessed"] == 1
        print("[PASS] Scenario 5: Performance upsert updated existing record without duplicate creation.")

        # -------------------------------------------------------------
        # SCENARIO 6: Duplicate Performance Prevention (Manual API Creation)
        # -------------------------------------------------------------
        dup_payload = {
            "schedule_id": str(sch_a1.schedule_id),
            "actual_start_date": "2026-03-01",
            "score": 4.0
        }
        res6 = client.post("/api/performance", json=dup_payload, headers=headers_a)
        assert res6.status_code == 409, f"Expected 409 Conflict, got {res6.status_code}: {res6.text}"
        assert "already exists for this schedule and start date" in res6.json()["detail"]
        print("[PASS] Scenario 6: Manual duplicate performance creation returned 409 Conflict.")

        # -------------------------------------------------------------
        # SCENARIO 7: Add Performance from Schedule Page / Endpoint
        # -------------------------------------------------------------
        sch_new_date_payload = {
            "actual_start_date": "2026-03-10",
            "actual_end_date": "2026-03-15",
            "score": 4.7,
            "feedback": "Created from Schedule page action"
        }
        res7 = client.post(f"/api/schedules/{sch_a1.schedule_id}/performance", json=sch_new_date_payload, headers=headers_a)
        assert res7.status_code == 201, f"Create from schedule failed: {res7.text}"
        perf_created7 = res7.json()
        assert perf_created7["schedule_id"] == str(sch_a1.schedule_id)
        assert perf_created7["score"] == 4.7
        print("[PASS] Scenario 7: Add Performance from Schedule endpoint succeeded.")

        # -------------------------------------------------------------
        # SCENARIO 8: Add Performance from Engineer Profile Endpoint
        # -------------------------------------------------------------
        prof_perf_payload = {
            "schedule_id": str(sch_a2.schedule_id),
            "orbit_id": orbit_id_a,
            "actual_start_date": "2026-04-10",
            "actual_end_date": "2026-04-18",
            "score": 4.9,
            "feedback": "Created from Engineer Profile page"
        }
        res8 = client.post("/api/performance", json=prof_perf_payload, headers=headers_a)
        assert res8.status_code == 201, f"Create from profile failed: {res8.text}"
        perf_created8 = res8.json()
        assert perf_created8["schedule_id"] == str(sch_a2.schedule_id)
        print("[PASS] Scenario 8: Add Performance from Engineer Profile endpoint succeeded.")

        # -------------------------------------------------------------
        # SCENARIO 9: Engineer Profile Schedule Filter
        # -------------------------------------------------------------
        res9 = client.get(f"/api/schedules?engineer_id={eng_a.engineer_id}", headers=headers_a)
        assert res9.status_code == 200
        items9 = res9.json()["items"]
        assert all(s["engineer_id"] == str(eng_a.engineer_id) for s in items9)
        assert len(items9) == 2
        print("[PASS] Scenario 9: Engineer Profile schedule selector returns only that engineer's schedules.")

        # -------------------------------------------------------------
        # SCENARIO 10: Score Range Validation (1.0 to 5.0)
        # -------------------------------------------------------------
        bad_score_payload = {
            "schedule_id": str(sch_a1.schedule_id),
            "actual_start_date": "2026-05-20",
            "score": 6.5
        }
        res10 = client.post("/api/performance", json=bad_score_payload, headers=headers_a)
        assert res10.status_code == 422
        print("[PASS] Scenario 10: Score > 5.0 rejected with HTTP 422.")

        # -------------------------------------------------------------
        # SCENARIO 11: Escalation Reason Validation
        # -------------------------------------------------------------
        bad_esc_payload = {
            "schedule_id": str(sch_a1.schedule_id),
            "actual_start_date": "2026-05-21",
            "score": 4.0,
            "escalation": True,
            "escalation_reason": ""
        }
        res11 = client.post("/api/performance", json=bad_esc_payload, headers=headers_a)
        assert res11.status_code == 422
        assert "Escalation reason is required" in res11.text
        print("[PASS] Scenario 11: Escalation without reason rejected with HTTP 422.")

        # -------------------------------------------------------------
        # SCENARIO 12: Date Order Validation
        # -------------------------------------------------------------
        bad_date_payload = {
            "schedule_id": str(sch_a1.schedule_id),
            "actual_start_date": "2026-05-20",
            "actual_end_date": "2026-05-10",
            "score": 4.0
        }
        res12 = client.post("/api/performance", json=bad_date_payload, headers=headers_a)
        assert res12.status_code == 422
        assert "actual_end_date should not be earlier" in res12.text
        print("[PASS] Scenario 12: Invalid date order rejected with HTTP 422.")

        # -------------------------------------------------------------
        # SCENARIO 13: Multi-tenant Isolation
        # -------------------------------------------------------------
        res13 = client.post(f"/api/schedules/{sch_b.schedule_id}/performance", json={"actual_start_date": "2026-03-05", "score": 4.0}, headers=headers_a)
        assert res13.status_code in (403, 404)
        print("[PASS] Scenario 13: Multi-tenant cross-company performance creation blocked.")

        print("\n=== ALL 13 PERFORMANCE V2 SCENARIOS PASSED SUCCESSFULLY! ===")

    finally:
        db.close()

if __name__ == "__main__":
    run_performance_v2_tests()
