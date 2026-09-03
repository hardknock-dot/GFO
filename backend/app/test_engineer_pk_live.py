import os
import sys
import io
import uuid
from datetime import datetime

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl

from fastapi.testclient import TestClient
from sqlalchemy import select

from app.main import app
from app.database import get_db, SessionLocal
from app.models.engineer import Engineer
from app.models.user import User
from app.models.company import Company

from app.services.security import create_access_token

client = TestClient(app)

def create_engineer_excel(headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Engineer"
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def run_live_engineer_pk_test():
    db = SessionLocal()
    try:
        # 1. Fetch real existing engineer from Supabase
        target_eng_id = uuid.UUID("19eab5e0-9ae9-4db2-9f1c-cdd8b2aa1d46")
        eng = db.get(Engineer, target_eng_id)
        if not eng:
            print(f"Engineer {target_eng_id} not found by fixed ID, searching first existing engineer...")
            eng = db.scalars(select(Engineer)).first()
            if not eng:
                print("ERROR: No engineers found in Supabase database!")
                sys.exit(1)
            target_eng_id = eng.engineer_id

        print("==================================================================")
        print("STAGE 1: REAL DATABASE INITIAL STATE")
        print(f"  Target engineer_id: {eng.engineer_id}")
        print(f"  Target engineer_name: {eng.engineer_name}")
        print(f"  Target employee_id: {eng.lam_id}")
        print(f"  Target orbit_id: {eng.orbit_id}")
        print(f"  Initial goes_by: '{eng.goes_by}'")
        print("==================================================================")

        initial_goes_by = eng.goes_by or "OriginalGoesBy"
        new_goes_by = f"DebugGoesBy_{datetime.now().strftime('%H%M%S')}"

        # Fetch company and user for auth
        company = db.get(Company, eng.company_id)
        admin_user = db.scalars(select(User).where(User.company_id == eng.company_id)).first()
        if not admin_user:
            admin_user = db.scalars(select(User)).first()

        token = create_access_token({"sub": str(admin_user.user_id)})
        headers_auth = {
            "Authorization": f"Bearer {token}",
            "X-Company-ID": str(company.company_id)
        }

        # 2. Construct Excel File
        excel_headers = [
            "engineer_id", "engineer_name", "goes_by", "customer_id", 
            "orbit_id", "level", "date_of_joining", "primary_tool", 
            "customer_experience", "industry_experience", "status", "email", "phone_number"
        ]

        excel_row = [
            str(eng.engineer_id),
            eng.engineer_name,
            new_goes_by,
            eng.lam_id or "74494",
            eng.orbit_id,
            eng.level or "L2 Specialist",
            str(eng.date_of_joining) if eng.date_of_joining else "2024-01-01",
            eng.primary_tool_type or "Tool A",
            float(eng.lam_experience) if eng.lam_experience is not None else 3.0,
            float(eng.industry_experience) if eng.industry_experience is not None else 5.0,
            eng.status or "Active",
            eng.email or "test@example.com",
            eng.phone_number or "+1234567890"
        ]

        print("\nSTAGE 2: EXCEL ROW CONSTRUCTED")
        print(f"  Excel engineer_id column: {excel_row[0]}")
        print(f"  Excel customer_id column: {excel_row[3]}")
        print(f"  Excel goes_by column: {excel_row[2]}")

        excel_bytes = create_engineer_excel(excel_headers, [excel_row])

        # 3. Perform Upload
        print("\nSTAGE 3: CALLING /api/upload ENDPOINT")
        response = client.post(
            "/api/upload",
            data={"module_id": "up-engineers"},
            files={"file": ("test_engineer.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=headers_auth
        )

        assert response.status_code == 200, f"Upload failed: {response.text}"
        res_data = response.json()

        print("\nSTAGE 4: ENDPOINT RESPONSE DATA")
        print(f"  inserted: {res_data.get('inserted')}")
        print(f"  updated: {res_data.get('updated')}")
        print(f"  unchanged: {res_data.get('unchanged')}")
        print(f"  errorsCount: {res_data.get('errorsCount')}")
        print(f"  message: {res_data.get('message')}")
        print(f"  report_url: {res_data.get('report_url')}")

        assert res_data["inserted"] == 0, f"Expected 0 inserted, got {res_data['inserted']}"
        assert res_data["updated"] == 1, f"Expected 1 updated, got {res_data['updated']}"
        assert res_data["errorsCount"] == 0, f"Expected 0 errors, got {res_data['errorsCount']}"

        # 4. Verify Database Persistence
        db.expire_all()
        updated_eng = db.get(Engineer, target_eng_id)
        
        print("\nSTAGE 5: DATABASE POST-UPDATE VERIFICATION")
        print(f"  Database engineer_id: {updated_eng.engineer_id} (Unchanged: {updated_eng.engineer_id == target_eng_id})")
        print(f"  Database employee_id (lam_id): {updated_eng.lam_id} (Unchanged: {updated_eng.lam_id == (eng.lam_id or '74494')})")
        print(f"  Database orbit_id: {updated_eng.orbit_id}")
        print(f"  Database goes_by: '{updated_eng.goes_by}' (Updated correctly: {updated_eng.goes_by == new_goes_by})")

        assert updated_eng.engineer_id == target_eng_id
        assert updated_eng.goes_by == new_goes_by

        # 5. Download and Verify Validation Report Workbook
        report_url = res_data["report_url"]
        report_filename = report_url.split("/")[-1]
        report_resp = client.get(f"/api/upload/download-report/{report_filename}", headers=headers_auth)
        assert report_resp.status_code == 200, f"Failed to download report: {report_resp.text}"

        report_wb = openpyxl.load_workbook(io.BytesIO(report_resp.content))
        print("\nSTAGE 6: VALIDATION WORKBOOK SHEETS VERIFICATION")
        print(f"  Sheet Names: {report_wb.sheetnames}")
        
        expected_sheets = ["Summary", "Valid Records", "Updated Records", "Unchanged Records", "Errors", "Duplicates"]
        for sheet_name in expected_sheets:
            assert sheet_name in report_wb.sheetnames, f"Missing expected sheet '{sheet_name}' in workbook!"

        ws_updated = report_wb["Updated Records"]
        updated_rows = list(ws_updated.iter_rows(values_only=True))
        print(f"\nSTAGE 7: UPDATED RECORDS SHEET DATA (Row count: {len(updated_rows)})")
        for i, row in enumerate(updated_rows):
            print(f"  Row {i+1}: {row}")

        assert len(updated_rows) >= 2, "Expected header + at least 1 updated row in 'Updated Records' sheet!"
        row_data = updated_rows[1]
        # Column 2 is Engineer ID
        report_eng_id = str(row_data[1])
        print(f"  Report Engineer ID in Updated Records: {report_eng_id}")
        assert report_eng_id == str(target_eng_id), f"Expected {target_eng_id}, got {report_eng_id}"

        # Clean up database goes_by to original if needed
        updated_eng.goes_by = eng.goes_by
        db.commit()

        print("\n==================================================================")
        print("ALL STAGE ASSERTIONS PASSED PERFECTLY!")
        print("==================================================================")

    finally:
        db.close()

if __name__ == "__main__":
    run_live_engineer_pk_test()
