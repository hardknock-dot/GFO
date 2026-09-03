import logging
import os
import re
import io
import time
import uuid as uuid_pkg
from datetime import datetime, date
from uuid import UUID
from typing import Optional, List, Dict, Any
from decimal import Decimal

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form, Header, Query
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import select, text, func

from app.database import get_db
from app.models.user import User
from app.models.company import Company
from app.models.engineer import Engineer
from app.models.skill import Skill
from app.models.schedule import Schedule
from app.models.visa import Visa
from app.models.travel import Travel
from app.models.performance import Performance
from app.models.leave import Leave
from app.services.auth_service import get_current_user, enforce_company_isolation, enforce_write_permission
from app.services.audit_service import log_audit
from app.services import bulk_upload_service
from app.schemas.bulk_upload import BulkUploadResponse
import openpyxl

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/upload", tags=["upload"], dependencies=[Depends(get_current_user)])

EMAIL_REGEX = re.compile(r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$")
TEMP_REPORTS_DIR = "backend/app/temp_reports"

def parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if not v:
        return None
    v_str = str(v).strip()
    if not v_str:
        return None
    if " " in v_str:
        v_str = v_str.split(" ")[0]
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(v_str, fmt).date()
        except ValueError:
            continue
    raise ValueError("Invalid date format")

def parse_experience(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    v_str = str(v).strip()
    if not v_str:
        return None
    v_clean = v_str.rstrip("+").rstrip(".").strip()
    match = re.match(r"^([\d.]+)\s*(?:years?|yrs?\.?)$", v_clean, re.IGNORECASE)
    if match:
        try:
            return float(match.group(1))
        except ValueError:
            pass
    try:
        return float(v_clean)
    except ValueError:
        raise ValueError("Must be a valid numeric experience value")

def clean_val(v):
    if v is None:
        return None
    if isinstance(v, str):
        v_stripped = v.strip()
        return v_stripped if v_stripped != "" else None
    return v

def normalize_header(name):
    if name is None:
        return ""
    return str(name).strip().lower().replace(" ", "").replace("_", "").replace("-", "").replace(".", "").replace("(", "").replace(")", "").replace("/", "").replace("#", "")

def norm_str(v):
    if v is None:
        return ""
    return str(v).strip().lower()

def norm_date(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        try:
            return parse_date(v).strftime("%Y-%m-%d")
        except Exception:
            return v.strip().lower()
    return str(v).strip().lower()

def norm_uuid(v):
    if v is None:
        return ""
    return str(v).strip().lower()

def parse_uuid_safe(val: Any) -> Optional[UUID]:
    if not val:
        return None
    val_str = str(val).strip()
    if not val_str:
        return None
    try:
        return UUID(val_str)
    except ValueError:
        raise ValueError(f"Invalid UUID format: '{val_str}'")

def values_are_equal(val1: Any, val2: Any) -> bool:
    """
    Type-normalized comparison helper to eliminate false-positive diffs.
    Handles None / empty strings, numbers (int, float, Decimal), dates, booleans, UUID strings.
    """
    if val1 is None and val2 is None:
        return True
    if val1 is None or val2 is None:
        if val1 is None and isinstance(val2, str) and val2.strip() == "":
            return True
        if val2 is None and isinstance(val1, str) and val1.strip() == "":
            return True
        return False

    # Dates / Datetimes
    if isinstance(val1, (date, datetime)) or isinstance(val2, (date, datetime)):
        try:
            d1 = val1.date() if isinstance(val1, datetime) else (parse_date(val1) if isinstance(val1, str) else val1)
            d2 = val2.date() if isinstance(val2, datetime) else (parse_date(val2) if isinstance(val2, str) else val2)
            return d1 == d2
        except Exception:
            pass

    # Numeric comparison (float, int, Decimal)
    if isinstance(val1, (int, float, Decimal)) or isinstance(val2, (int, float, Decimal)):
        try:
            n1 = float(val1)
            n2 = float(val2)
            return abs(n1 - n2) < 1e-5
        except (ValueError, TypeError):
            pass

    # Boolean comparison
    if isinstance(val1, bool) or isinstance(val2, bool):
        try:
            b1 = parse_boolean(val1) if not isinstance(val1, bool) else val1
            b2 = parse_boolean(val2) if not isinstance(val2, bool) else val2
            return b1 == b2
        except Exception:
            pass

    # String comparison
    str1 = str(val1).strip()
    str2 = str(val2).strip()
    return str1 == str2

HEADER_MAP = {
    "engineerid": "engineer_id",
    "engineer_id": "engineer_id",
    "id": "engineer_id",
    "engineername": "engineer_name",
    "name": "engineer_name",
    "engineer": "engineer_name",
    "goesby": "goes_by",
    "preferredname": "goes_by",
    "customerid": "employee_id",
    "customer_id": "employee_id",
    "custid": "employee_id",
    "cust_id": "employee_id",
    "lamid": "employee_id",
    "lam_id": "employee_id",
    "employeeid": "employee_id",
    "employee_id": "employee_id",
    "empid": "employee_id",
    "emp_id": "employee_id",
    "employeenumber": "employee_id",
    "employee_number": "employee_id",
    "employeeno": "employee_id",
    "employee_no": "employee_id",
    "empno": "employee_id",
    "emp_no": "employee_id",
    "customernumber": "employee_id",
    "customer_number": "employee_id",
    "customerno": "employee_id",
    "customer_no": "employee_id",
    "custno": "employee_id",
    "cust_no": "employee_id",
    "employeecode": "employee_id",
    "employee_code": "employee_id",
    "empcode": "employee_id",
    "emp_code": "employee_id",
    "staffid": "employee_id",
    "staff_id": "employee_id",
    "orbitid": "orbit_id",
    "orbit": "orbit_id",
    "level": "level",
    "engineerlevel": "level",
    "dateofjoining": "date_of_joining",
    "joiningdate": "date_of_joining",
    "doj": "date_of_joining",
    "primarytooltype": "primary_tool",
    "primarytool": "primary_tool",
    "tooltype": "primary_tool",
    "tool": "primary_tool",
    "lamexperience": "customer_experience",
    "customerexperience": "customer_experience",
    "industryexperience": "industry_experience",
    "status": "status",
    "engineerstatus": "status",
    "email": "email",
    "emailaddress": "email",
    "phonenumber": "phone_number",
    "phone": "phone_number",
    "mobile": "phone_number",
    "contact": "phone_number",
    "contactnumber": "phone_number"
}

def map_engineer_header(raw_header: str) -> Optional[str]:
    if not raw_header:
        return None
    norm = normalize_header(raw_header)
    
    if norm in HEADER_MAP:
        return HEADER_MAP[norm]

    if norm in ("engineerid", "id", "engineer_id"):
        return "engineer_id"
        
    if ("customer" in norm or "lam" in norm or "cust" in norm) and ("exp" in norm or "experience" in norm):
        return "customer_experience"
        
    if ("industry" in norm or "ind" in norm) and ("exp" in norm or "experience" in norm):
        return "industry_experience"
        
    if "date" in norm and ("join" in norm or "doj" in norm):
        return "date_of_joining"
        
    if "orbit" in norm:
        return "orbit_id"
        
    if "email" in norm:
        return "email"
        
    if "phone" in norm or "mobile" in norm or "contact" in norm:
        return "phone_number"
        
    if "level" in norm:
        return "level"
        
    if "status" in norm:
        return "status"
        
    if "tool" in norm:
        return "primary_tool"
        
    if ("employee" in norm or "emp" in norm or "lam" in norm or "customer" in norm or "cust" in norm or "staff" in norm) and ("id" in norm or "num" in norm or "number" in norm or "no" in norm or "code" in norm):
        return "employee_id"
        
    if "name" in norm and "engineer" in norm:
        return "engineer_name"
        
    return None

def parse_boolean(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        if int(v) == 1:
            return True
        if int(v) == 0:
            return False
        raise ValueError("Invalid boolean integer value")
    v_str = str(v).strip().lower()
    if not v_str:
        return None
    if v_str in ("yes", "y", "true", "t", "1"):
        return True
    if v_str in ("no", "n", "false", "f", "0"):
        return False
    raise ValueError("Invalid boolean representation")

SKILL_HEADER_MAP = {
    "skillid": "skill_id",
    "skill_id": "skill_id",
    "id": "skill_id",
    "orbitid": "orbit_id",
    "country": "country",
    "fab": "fab",
    "wafersize": "wafer_size",
    "tooltype": "tool_type",
    "startdate": "start_date",
    "enddate": "end_date",
    "#oftools": "number_of_tools",
    "numberoftools": "number_of_tools",
    "numoftools": "number_of_tools",
    "oftools": "number_of_tools",
    "role": "role",
    "previousprocessstartupexperience": "previous_process_startup",
    "previousprocessstartup": "previous_process_startup",
    "previouscmpmexperience": "previous_cm_pm",
    "previouscmpm": "previous_cm_pm",
    "previouscm/pmexperience": "previous_cm_pm",
    "previouscm/pm": "previous_cm_pm",
    "readyforprimaryrole": "ready_for_primary_role",
    "comments": "comments"
}

SCHEDULE_HEADER_MAP = {
    "scheduleid": "schedule_id",
    "schedule_id": "schedule_id",
    "id": "schedule_id",
    "orbitid": "orbit_id",
    "supporttype": "support_type",
    "country": "country",
    "fabcity": "fab_city",
    "fabsite": "fab_site",
    "startdate": "start_date",
    "enddate": "end_date",
    "schedulestatus": "schedule_status",
    "remarks": "remarks",
    "commentstatus": "comment_status",
    "owner": "owner",
    "ownerid": "owner"
}

VISA_HEADER_MAP = {
    "visaid": "visa_id",
    "visa_id": "visa_id",
    "id": "visa_id",
    "orbitid": "orbit_id",
    "country": "country",
    "visatype": "visa_type",
    "type": "visa_type",
    "appliedon": "applied_on",
    "applieddate": "applied_on",
    "applied": "applied_on",
    "startdate": "visa_start_date",
    "visastartdate": "visa_start_date",
    "issuedate": "visa_start_date",
    "start": "visa_start_date",
    "enddate": "visa_end_date",
    "visaenddate": "visa_end_date",
    "expirydate": "visa_end_date",
    "expirationdate": "visa_end_date",
    "end": "visa_end_date",
    "comments": "comments",
    "remarks": "comments",
    "owner": "owner",
    "owneremail": "owner",
    "ownername": "owner",
    "commentstatus": "comment_status"
}

TRAVEL_HEADER_MAP = {
    "travelid": "travel_id",
    "travel_id": "travel_id",
    "id": "travel_id",
    "orbitid": "orbit_id",
    "engineername": "engineer_name",
    "bookingdate": "booking_date",
    "traveldate": "travel_date",
    "departuredate": "travel_date",
    "flightdate": "travel_date",
    "purpose": "purpose",
    "travelpurpose": "purpose",
    "comments": "comments",
    "remarks": "comments",
    "notes": "comments",
    "supporttype": "support_type",
    "country": "country",
    "fabcity": "fab_city",
    "fabsite": "fab_site",
    "scheduleid": "schedule_id",
    "owner": "owner",
    "ownerid": "owner"
}

PERFORMANCE_HEADER_MAP = {
    "performanceid": "performance_id",
    "performance_id": "performance_id",
    "id": "performance_id",
    "orbitid": "orbit_id",
    "engineername": "engineer_name",
    "actualstartdate": "actual_start_date",
    "actualstart": "actual_start_date",
    "startdate": "actual_start_date",
    "actualenddate": "actual_end_date",
    "actualend": "actual_end_date",
    "enddate": "actual_end_date",
    "escalation": "escalation",
    "escalated": "escalation",
    "escalationreason": "escalation_reason",
    "reason": "escalation_reason",
    "feedback": "feedback",
    "notes": "feedback",
    "comments": "feedback",
    "score": "score",
    "rating": "score",
    "attachment": "attachment",
    "supporttype": "support_type",
    "country": "country",
    "fabcity": "fab_city",
    "fabsite": "fab_site",
    "scheduleid": "schedule_id",
    "owner": "owner",
    "ownerid": "owner"
}

LEAVE_HEADER_MAP = {
    "leaveid": "leave_id",
    "leave_id": "leave_id",
    "id": "leave_id",
    "orbitid": "orbit_id",
    "engineername": "engineer_name",
    "leavetype": "leave_type",
    "type": "leave_type",
    "category": "leave_type",
    "requesteddate": "requested_date",
    "requested": "requested_date",
    "date": "requested_date",
    "absencedate": "requested_date",
    "startdate": "requested_date",
    "leavedate": "requested_date",
    "requestedon": "requested_on",
    "submissiondate": "requested_on",
    "submittedon": "requested_on",
    "submitted": "requested_on",
    "approvalstatus": "approval_status",
    "status": "approval_status",
    "owner": "owner",
    "ownerid": "owner"
}

@router.post("")
async def bulk_upload(
    file: UploadFile = File(...),
    module_id: str = Form(...),
    x_company_id: Optional[str] = Header(None, alias="X-Company-ID"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Process Excel/CSV file bulk upload.
    Validate rows against requirements, insert valid new rows into the database,
    update existing rows by primary key, and generate a validation report file.
    """
    enforce_write_permission(current_user)

    target_company_id = None
    if x_company_id:
        try:
            target_company_id = UUID(x_company_id)
        except ValueError:
            comp = db.scalars(
                select(Company).where(
                    (Company.company_code == x_company_id) | 
                    (Company.company_name.ilike(x_company_id))
                )
            ).first()
            if comp:
                target_company_id = comp.company_id

    if current_user.role != 'Global Admin':
        target_company_id = current_user.company_id
    elif target_company_id is None:
        comp = db.scalars(select(Company)).first()
        if comp:
            target_company_id = comp.company_id

    company = db.get(Company, target_company_id)
    if not company:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Target company tenant not found."
        )

    enforce_company_isolation(db, current_user, target_company_id)

    upload_type = "engineers"
    if module_id == "up-skills":
        upload_type = "skills"
    elif module_id == "up-schedule":
        upload_type = "schedules"
    elif module_id == "up-visa":
        upload_type = "visas"
    elif module_id == "up-travel":
        upload_type = "travel"
    elif module_id == "up-performance":
        upload_type = "performance"
    elif module_id == "up-leave":
        upload_type = "leaves"
    elif module_id != "up-engineers":
        upload_type = module_id

    db_upload = bulk_upload_service.create_bulk_upload(
        db,
        company_id=target_company_id,
        uploaded_by=current_user.user_id,
        file_name=file.filename,
        upload_type=upload_type
    )
    upload_id = db_upload.upload_id

    log_audit(
        db=db,
        user_id=current_user.user_id,
        company_id=target_company_id,
        action="BULK_UPLOAD",
        entity_type="BulkUpload",
        entity_id=upload_id,
        description=f"Bulk upload performed for {upload_type}: {file.filename}"
    )

    try:
        if module_id not in ("up-engineers", "up-skills", "up-schedule", "up-visa", "up-travel", "up-performance", "up-leave"):
            bulk_upload_service.update_bulk_upload(
                db,
                upload_id=upload_id,
                status="COMPLETED",
                total_rows=0,
                valid_rows=0,
                error_rows=0
            )
            return {
                "success": True,
                "rowsProcessed": 0,
                "errorsCount": 0,
                "message": f"Upload validation succeeded. Module {module_id} is currently in read-only validation state."
            }

        # User lookup helper for owner resolution
        company_users = db.scalars(
            select(User).where(User.company_id == target_company_id)
        ).all()
        user_by_email = {u.email.lower(): u.user_id for u in company_users if u.email}
        user_by_name = {u.full_name.lower(): u.user_id for u in company_users if u.full_name}
        user_by_id = {str(u.user_id): u.user_id for u in company_users}

        def resolve_owner(owner_raw):
            if owner_raw is not None and str(owner_raw).strip() != "":
                clean_owner = str(owner_raw).strip().lower()
                return user_by_email.get(clean_owner) or user_by_name.get(clean_owner) or user_by_id.get(str(owner_raw).strip())
            return None

        # =========================================================================
        # MODULE 1: UP-SKILLS
        # =========================================================================
        if module_id == "up-skills":
            start_time = time.perf_counter()
            try:
                contents = await file.read()
                wb = openpyxl.load_workbook(io.BytesIO(contents))
            except Exception:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Failed to parse Excel file. Please ensure it is a valid .xlsx file."
                )

            skill_sheet_name = None
            for name in wb.sheetnames:
                if name.strip().lower() in ("skill matrix", "skills", "skill"):
                    skill_sheet_name = name
                    break

            if not skill_sheet_name:
                if len(wb.sheetnames) == 1:
                    skill_sheet_name = wb.sheetnames[0]
                else:
                    raise HTTPException(
                        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                        detail="Excel workbook must contain a Skill Matrix sheet."
                    )

            sheet = wb[skill_sheet_name]
            first_row = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
            col_indices = {}
            for idx, val in enumerate(first_row):
                if val is not None:
                    norm = normalize_header(val)
                    mapped_field = SKILL_HEADER_MAP.get(norm)
                    if mapped_field:
                        col_indices[mapped_field] = idx + 1

            raw_rows = []
            for r in range(2, sheet.max_row + 1):
                is_blank = True
                for c in range(1, sheet.max_column + 1):
                    val = sheet.cell(row=r, column=c).value
                    if val is not None and str(val).strip() != "":
                        is_blank = False
                        break
                if is_blank:
                    continue

                row_dict = {"excel_row": r}
                original_engineer_name = None
                for idx, val in enumerate(first_row):
                    if val is not None:
                        norm = normalize_header(val)
                        if norm in ("engineername", "name"):
                            original_engineer_name = clean_val(sheet.cell(row=r, column=idx + 1).value)
                            break
                row_dict["original_engineer_name"] = original_engineer_name

                for field, col_idx in col_indices.items():
                    row_dict[field] = clean_val(sheet.cell(row=r, column=col_idx).value)

                raw_rows.append(row_dict)

            if not raw_rows:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="The Skill Matrix sheet is empty or contains no rows."
                )

            unique_orbit_ids = {
                norm_str(row.get("orbit_id"))
                for row in raw_rows
                if row.get("orbit_id") and norm_str(row.get("orbit_id")) != ""
            }

            db_engineers = []
            if unique_orbit_ids:
                db_engineers = db.scalars(
                    select(Engineer).where(
                        func.lower(Engineer.orbit_id).in_(list(unique_orbit_ids)),
                        Engineer.company_id == target_company_id
                    )
                ).all()

            orbit_to_engineer = {
                norm_str(eng.orbit_id): (eng.engineer_id, eng.engineer_name)
                for eng in db_engineers
            }

            errors_list = []
            duplicates_list = []
            existing_list = []
            unchanged_list = []
            valid_rows_to_insert = []
            seen_keys = set()

            total_rows = len(raw_rows)

            for row_dict in raw_rows:
                row_errors = []
                raw_pk = row_dict.get("skill_id")
                parsed_pk = None

                if raw_pk is not None and str(raw_pk).strip() != "":
                    try:
                        parsed_pk = parse_uuid_safe(raw_pk)
                    except ValueError as ve:
                        row_errors.append({"field": "Skill ID", "value": str(raw_pk), "error": str(ve)})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                # Parse common fields
                start_date = None
                if row_dict.get("start_date") is not None:
                    try:
                        start_date = parse_date(row_dict["start_date"])
                    except ValueError:
                        row_errors.append({"field": "Start Date", "value": str(row_dict["start_date"]), "error": "Invalid start date format."})

                end_date = None
                if row_dict.get("end_date") is not None:
                    try:
                        end_date = parse_date(row_dict["end_date"])
                    except ValueError:
                        row_errors.append({"field": "End Date", "value": str(row_dict["end_date"]), "error": "Invalid end date format."})

                if start_date and end_date and end_date < start_date:
                    row_errors.append({"field": "End Date", "value": str(row_dict["end_date"]), "error": "End date should not be earlier than start_date"})

                number_of_tools = None
                if row_dict.get("number_of_tools") is not None:
                    try:
                        num_val = float(row_dict["number_of_tools"])
                        if not num_val.is_integer() or num_val < 0:
                            raise ValueError()
                        number_of_tools = int(num_val)
                    except (ValueError, TypeError):
                        row_errors.append({"field": "# of Tools", "value": str(row_dict["number_of_tools"]), "error": "Number of Tools must be a non-negative integer."})

                previous_process_startup = None
                if row_dict.get("previous_process_startup") is not None:
                    try:
                        previous_process_startup = parse_boolean(row_dict["previous_process_startup"])
                    except ValueError:
                        row_errors.append({"field": "Previous Process Startup Experience", "value": str(row_dict["previous_process_startup"]), "error": "Invalid boolean value representation."})

                previous_cm_pm = None
                if row_dict.get("previous_cm_pm") is not None:
                    try:
                        previous_cm_pm = parse_boolean(row_dict["previous_cm_pm"])
                    except ValueError:
                        row_errors.append({"field": "Previous CM/PM Experience", "value": str(row_dict["previous_cm_pm"]), "error": "Invalid boolean value representation."})

                ready_for_primary_role = None
                if row_dict.get("ready_for_primary_role") is not None:
                    try:
                        ready_for_primary_role = parse_boolean(row_dict["ready_for_primary_role"])
                    except ValueError:
                        row_errors.append({"field": "Ready for Primary Role", "value": str(row_dict["ready_for_primary_role"]), "error": "Invalid boolean value representation."})

                row_dict["start_date"] = start_date
                row_dict["end_date"] = end_date
                row_dict["number_of_tools"] = number_of_tools
                row_dict["previous_process_startup"] = previous_process_startup
                row_dict["previous_cm_pm"] = previous_cm_pm
                row_dict["ready_for_primary_role"] = ready_for_primary_role

                if row_errors:
                    row_dict["errors"] = row_errors
                    errors_list.append(row_dict)
                    continue

                # PRIMARY KEY UPDATE PATH
                if parsed_pk is not None:
                    db_skill = db.get(Skill, parsed_pk)
                    if not db_skill:
                        row_errors.append({"field": "Skill ID", "value": str(parsed_pk), "error": f"Record with Skill ID '{parsed_pk}' was not found. Cannot update non-existent record ID."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    # Verify Tenant
                    eng = db.get(Engineer, db_skill.engineer_id)
                    if not eng or eng.company_id != target_company_id:
                        row_errors.append({"field": "Skill ID", "value": str(parsed_pk), "error": f"Record with Skill ID '{parsed_pk}' belongs to another company."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    row_dict["resolved_engineer_name"] = eng.engineer_name
                    row_dict["orbit_id"] = eng.orbit_id

                    changes = []
                    field_specs = [
                        ("country", "Country", db_skill.country),
                        ("fab", "Fab", db_skill.fab),
                        ("wafer_size", "Wafer Size", db_skill.wafer_size),
                        ("tool_type", "Tool Type", db_skill.tool_type),
                        ("start_date", "Start Date", db_skill.start_date),
                        ("end_date", "End Date", db_skill.end_date),
                        ("number_of_tools", "Number of Tools", db_skill.number_of_tools),
                        ("role", "Role", db_skill.role),
                        ("previous_process_startup", "Previous Process Startup", db_skill.previous_process_startup),
                        ("previous_cm_pm", "Previous CM/PM", db_skill.previous_cm_pm),
                        ("ready_for_primary_role", "Ready for Primary Role", db_skill.ready_for_primary_role),
                        ("comments", "Comments", db_skill.comments)
                    ]

                    for f_key, f_label, cur_val in field_specs:
                        if f_key in col_indices:
                            new_val = row_dict.get(f_key)
                            if not values_are_equal(cur_val, new_val):
                                changes.append(f"{f_label}: '{cur_val}' -> '{new_val}'")
                                setattr(db_skill, f_key, new_val)

                    if changes:
                        db_skill.updated_at = datetime.utcnow()
                        row_dict["update_status"] = "UPDATED"
                        row_dict["changed_fields"] = "; ".join(changes)
                        existing_list.append(row_dict)
                    else:
                        row_dict["update_status"] = "UNCHANGED"
                        row_dict["changed_fields"] = "No fields modified"
                        unchanged_list.append(row_dict)

                # NEW RECORD INSERTION PATH (No Primary Key)
                else:
                    orbit_id = row_dict.get("orbit_id")
                    if not orbit_id:
                        row_errors.append({"field": "Orbit ID", "value": "", "error": "Orbit ID is required for new record creation."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    eng_info = orbit_to_engineer.get(norm_str(orbit_id))
                    if not eng_info:
                        row_errors.append({"field": "Orbit ID", "value": str(orbit_id), "error": f"Engineer with Orbit ID '{orbit_id}' does not exist in the selected company."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    engineer_id, resolved_engineer_name = eng_info
                    row_dict["engineer_id"] = engineer_id
                    row_dict["resolved_engineer_name"] = resolved_engineer_name

                    country = row_dict.get("country")
                    fab = row_dict.get("fab")
                    wafer_size = row_dict.get("wafer_size")
                    tool_type = row_dict.get("tool_type")

                    row_key = (norm_uuid(engineer_id), norm_str(country), norm_str(fab), norm_str(wafer_size), norm_str(tool_type), norm_date(start_date), norm_date(end_date))
                    if row_key in seen_keys:
                        row_dict["duplicate_key"] = f"EngineerID: {engineer_id}, Tool: {tool_type}, Fab: {fab}"
                        duplicates_list.append(row_dict)
                        continue
                    seen_keys.add(row_key)

                    valid_rows_to_insert.append(row_dict)

            # Persist Inserts & Updates
            imported_count = 0
            failed_count = 0
            try:
                skills_to_add = []
                for item in valid_rows_to_insert:
                    db_skill = Skill(
                        skill_id=uuid_pkg.uuid4(),
                        engineer_id=item["engineer_id"],
                        country=item.get("country"),
                        fab=item.get("fab"),
                        wafer_size=item.get("wafer_size"),
                        tool_type=item.get("tool_type"),
                        start_date=item.get("start_date"),
                        end_date=item.get("end_date"),
                        number_of_tools=item.get("number_of_tools"),
                        role=item.get("role"),
                        previous_process_startup=item.get("previous_process_startup"),
                        previous_cm_pm=item.get("previous_cm_pm"),
                        ready_for_primary_role=item.get("ready_for_primary_role"),
                        comments=item.get("comments"),
                        created_at=datetime.utcnow(),
                        updated_at=datetime.utcnow()
                    )
                    skills_to_add.append(db_skill)
                if skills_to_add:
                    db.add_all(skills_to_add)
                db.commit()
                imported_count = len(valid_rows_to_insert) + len(existing_list)
            except Exception as insert_err:
                db.rollback()
                failed_count = len(valid_rows_to_insert) + len(existing_list)
                bulk_upload_service.update_bulk_upload(
                    db,
                    upload_id=upload_id,
                    status="FAILED",
                    failed_rows=failed_count
                )
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail=f"Database ingestion failed: {str(insert_err)}"
                )

            # Generate Report
            report_wb = openpyxl.Workbook()
            ws_summary = report_wb.active
            ws_summary.title = "Summary"
            ws_summary.append(["ORMP Skill Matrix Bulk Ingestion Report"])
            ws_summary.append([])
            ws_summary.append(["File Name", file.filename])
            ws_summary.append(["Uploaded By", current_user.full_name])
            ws_summary.append(["Upload Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            ws_summary.append(["Target Company", company.company_name])
            ws_summary.append(["Company UUID", str(company.company_id)])
            ws_summary.append([])
            ws_summary.append(["Metric", "Count"])
            ws_summary.append(["Total Rows", total_rows])
            ws_summary.append(["Inserted Records", len(valid_rows_to_insert)])
            ws_summary.append(["Updated Records", len(existing_list)])
            ws_summary.append(["Unchanged Records", len(unchanged_list)])
            ws_summary.append(["Error Rows", len(errors_list)])
            ws_summary.append(["Duplicate Rows", len(duplicates_list)])
            ws_summary.append(["Warning Rows", 0])

            for col in ws_summary.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws_summary.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)

            headers_valid = [
                "Excel Row", "Skill ID", "Orbit ID", "Engineer Name", "Country", "Fab", "Wafer Size", 
                "Tool Type", "Start Date", "End Date", "Number of Tools", "Role", 
                "Previous Process Startup", "Previous CM/PM", "Ready for Primary Role", "Comments", "Status"
            ]

            ws_valid = report_wb.create_sheet(title="Valid Records")
            ws_valid.append(headers_valid)
            for r in valid_rows_to_insert:
                ws_valid.append([
                    r["excel_row"],
                    str(r.get("skill_id") or ""),
                    r.get("orbit_id"),
                    r.get("resolved_engineer_name"),
                    r.get("country"),
                    r.get("fab"),
                    r.get("wafer_size"),
                    r.get("tool_type"),
                    str(r.get("start_date")) if r.get("start_date") else "",
                    str(r.get("end_date")) if r.get("end_date") else "",
                    r.get("number_of_tools"),
                    r.get("role"),
                    "Yes" if r.get("previous_process_startup") else ("No" if r.get("previous_process_startup") is False else ""),
                    "Yes" if r.get("previous_cm_pm") else ("No" if r.get("previous_cm_pm") is False else ""),
                    "Yes" if r.get("ready_for_primary_role") else ("No" if r.get("ready_for_primary_role") is False else ""),
                    r.get("comments"),
                    "INSERTED"
                ])

            ws_updated = report_wb.create_sheet(title="Updated Records")
            ws_updated.append(["Excel Row", "Skill ID", "Orbit ID", "Engineer Name", "Action Status", "Changed Columns", "Country", "Fab", "Tool Type", "Role", "Comments"])
            for r in existing_list:
                ws_updated.append([
                    r["excel_row"],
                    str(r.get("skill_id") or ""),
                    r.get("orbit_id") or "",
                    r.get("resolved_engineer_name") or "",
                    "UPDATED",
                    r.get("changed_fields") or "",
                    r.get("country"),
                    r.get("fab"),
                    r.get("tool_type"),
                    r.get("role"),
                    r.get("comments")
                ])

            ws_unchanged = report_wb.create_sheet(title="Unchanged Records")
            ws_unchanged.append(["Excel Row", "Skill ID", "Orbit ID", "Engineer Name", "Action Status", "Details", "Country", "Fab", "Tool Type", "Role", "Comments"])
            for r in unchanged_list:
                ws_unchanged.append([
                    r["excel_row"],
                    str(r.get("skill_id") or ""),
                    r.get("orbit_id") or "",
                    r.get("resolved_engineer_name") or "",
                    "UNCHANGED",
                    "All supplied values match database",
                    r.get("country"),
                    r.get("fab"),
                    r.get("tool_type"),
                    r.get("role"),
                    r.get("comments")
                ])

            ws_errors = report_wb.create_sheet(title="Errors")
            ws_errors.append(["Excel Row", "Orbit ID", "Field", "Value", "Error"])
            for r in errors_list:
                for err in r.get("errors", []):
                    ws_errors.append([
                        r["excel_row"],
                        r.get("orbit_id") or "",
                        err.get("field") or "",
                        err.get("value") or "",
                        err.get("error") or ""
                    ])

            ws_dups = report_wb.create_sheet(title="Duplicates")
            ws_dups.append(["Excel Row", "Orbit ID", "Duplicate Key", "Duplicate Rows", "Reason"])
            for r in duplicates_list:
                ws_dups.append([
                    r["excel_row"],
                    r.get("orbit_id") or "",
                    r.get("duplicate_key") or "",
                    "",
                    "Duplicate Skill row in Excel sheet"
                ])

            ws_warn = report_wb.create_sheet(title="Warnings")
            ws_warn.append(["Excel Row", "Orbit ID", "Field", "Value", "Warning"])

            for sheet_obj in (ws_valid, ws_updated, ws_unchanged, ws_errors, ws_dups, ws_warn):
                for col in sheet_obj.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    sheet_obj.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)

            os.makedirs(TEMP_REPORTS_DIR, exist_ok=True)
            report_filename = f"validation_report_{uuid_pkg.uuid4()}.xlsx"
            report_path = os.path.join(TEMP_REPORTS_DIR, report_filename)
            report_wb.save(report_path)

            final_status = "COMPLETED"
            if len(errors_list) > 0 or len(duplicates_list) > 0:
                final_status = "COMPLETED_WITH_ERRORS"

            bulk_upload_service.update_bulk_upload(
                db,
                upload_id=upload_id,
                status=final_status,
                report_file=report_filename,
                imported_rows=imported_count,
                failed_rows=failed_count
            )

            ingested_msg = f"Ingested {len(valid_rows_to_insert)} new skill records. Updated {len(existing_list)} existing skill records. {len(unchanged_list)} records were unchanged."
            if errors_list or duplicates_list:
                ingested_msg += " Some rows had validation errors or duplicates. See the validation report for details."

            return {
                "success": True,
                "rowsProcessed": total_rows,
                "errorsCount": len(errors_list),
                "inserted": len(valid_rows_to_insert),
                "updated": len(existing_list),
                "unchanged": len(unchanged_list),
                "message": ingested_msg,
                "report_url": f"/api/upload/download-report/{report_filename}"
            }

        # =========================================================================
        # MODULE 2: UP-SCHEDULE
        # =========================================================================
        if module_id == "up-schedule":
            start_time = time.perf_counter()
            try:
                contents = await file.read()
                wb = openpyxl.load_workbook(io.BytesIO(contents))
            except Exception:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Failed to parse Excel file. Please ensure it is a valid .xlsx file."
                )

            schedule_sheet_name = None
            for name in wb.sheetnames:
                if name.strip().lower() in ("schedule", "schedules"):
                    schedule_sheet_name = name
                    break

            if not schedule_sheet_name:
                if len(wb.sheetnames) == 1:
                    schedule_sheet_name = wb.sheetnames[0]
                else:
                    raise HTTPException(
                        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                        detail="Excel workbook must contain a Schedule sheet."
                    )

            sheet = wb[schedule_sheet_name]
            first_row = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
            col_indices = {}
            for idx, val in enumerate(first_row):
                if val is not None:
                    norm = normalize_header(val)
                    mapped_field = SCHEDULE_HEADER_MAP.get(norm)
                    if mapped_field:
                        col_indices[mapped_field] = idx + 1

            raw_rows = []
            for r in range(2, sheet.max_row + 1):
                is_blank = True
                for c in range(1, sheet.max_column + 1):
                    val = sheet.cell(row=r, column=c).value
                    if val is not None and str(val).strip() != "":
                        is_blank = False
                        break
                if is_blank:
                    continue

                row_dict = {"excel_row": r}
                original_engineer_name = None
                for idx, val in enumerate(first_row):
                    if val is not None:
                        norm = normalize_header(val)
                        if norm in ("engineername", "name"):
                            original_engineer_name = clean_val(sheet.cell(row=r, column=idx + 1).value)
                            break
                row_dict["original_engineer_name"] = original_engineer_name

                for field, col_idx in col_indices.items():
                    row_dict[field] = clean_val(sheet.cell(row=r, column=col_idx).value)

                raw_rows.append(row_dict)

            if not raw_rows:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="The Schedule sheet is empty or contains no rows."
                )

            unique_orbit_ids = {
                norm_str(row.get("orbit_id"))
                for row in raw_rows
                if row.get("orbit_id") and norm_str(row.get("orbit_id")) != ""
            }

            db_engineers = []
            if unique_orbit_ids:
                db_engineers = db.scalars(
                    select(Engineer).where(
                        func.lower(Engineer.orbit_id).in_(list(unique_orbit_ids)),
                        Engineer.company_id == target_company_id
                    )
                ).all()

            orbit_to_engineer = {
                norm_str(eng.orbit_id): (eng.engineer_id, eng.engineer_name)
                for eng in db_engineers
            }

            errors_list = []
            duplicates_list = []
            existing_list = []
            unchanged_list = []
            valid_rows_to_insert = []
            seen_keys = set()

            total_rows = len(raw_rows)

            for row_dict in raw_rows:
                row_errors = []
                raw_pk = row_dict.get("schedule_id")
                parsed_pk = None

                if raw_pk is not None and str(raw_pk).strip() != "":
                    try:
                        parsed_pk = parse_uuid_safe(raw_pk)
                    except ValueError as ve:
                        row_errors.append({"field": "Schedule ID", "value": str(raw_pk), "error": str(ve)})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                start_date = None
                if row_dict.get("start_date") is not None:
                    try:
                        start_date = parse_date(row_dict["start_date"])
                    except ValueError:
                        row_errors.append({"field": "Start Date", "value": str(row_dict["start_date"]), "error": "Invalid start date format."})

                end_date = None
                if row_dict.get("end_date") is not None:
                    try:
                        end_date = parse_date(row_dict["end_date"])
                    except ValueError:
                        row_errors.append({"field": "End Date", "value": str(row_dict["end_date"]), "error": "Invalid end date format."})

                if start_date and end_date and end_date < start_date:
                    row_errors.append({"field": "End Date", "value": str(row_dict["end_date"]), "error": "End date should not be earlier than start_date"})

                resolved_owner_id = resolve_owner(row_dict.get("owner"))
                row_dict["owner_id"] = resolved_owner_id
                row_dict["start_date"] = start_date
                row_dict["end_date"] = end_date

                if row_errors:
                    row_dict["errors"] = row_errors
                    errors_list.append(row_dict)
                    continue

                # PRIMARY KEY UPDATE PATH
                if parsed_pk is not None:
                    db_sched = db.get(Schedule, parsed_pk)
                    if not db_sched:
                        row_errors.append({"field": "Schedule ID", "value": str(parsed_pk), "error": f"Record with Schedule ID '{parsed_pk}' was not found. Cannot update non-existent record ID."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    eng = db.get(Engineer, db_sched.engineer_id)
                    if not eng or eng.company_id != target_company_id:
                        row_errors.append({"field": "Schedule ID", "value": str(parsed_pk), "error": f"Record with Schedule ID '{parsed_pk}' belongs to another company."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    row_dict["resolved_engineer_name"] = eng.engineer_name
                    row_dict["orbit_id"] = eng.orbit_id

                    changes = []
                    field_specs = [
                        ("support_type", "Support Type", db_sched.support_type),
                        ("country", "Country", db_sched.country),
                        ("fab_city", "Fab City", db_sched.fab_city),
                        ("fab_site", "Fab Site", db_sched.fab_site),
                        ("start_date", "Start Date", db_sched.start_date),
                        ("end_date", "End Date", db_sched.end_date),
                        ("schedule_status", "Schedule Status", db_sched.schedule_status),
                        ("remarks", "Remarks", db_sched.remarks),
                        ("comment_status", "Comment Status", db_sched.comment_status),
                        ("owner_id", "Owner ID", db_sched.owner_id)
                    ]

                    for f_key, f_label, cur_val in field_specs:
                        if f_key in col_indices or (f_key == "owner_id" and "owner" in col_indices):
                            new_val = row_dict.get(f_key)
                            if not values_are_equal(cur_val, new_val):
                                changes.append(f"{f_label}: '{cur_val}' -> '{new_val}'")
                                setattr(db_sched, f_key, new_val)

                    if changes:
                        db_sched.updated_at = datetime.utcnow()
                        row_dict["update_status"] = "UPDATED"
                        row_dict["changed_fields"] = "; ".join(changes)
                        existing_list.append(row_dict)
                    else:
                        row_dict["update_status"] = "UNCHANGED"
                        row_dict["changed_fields"] = "No fields modified"
                        unchanged_list.append(row_dict)

                # NEW RECORD INSERTION PATH
                else:
                    orbit_id = row_dict.get("orbit_id")
                    if not orbit_id:
                        row_errors.append({"field": "Orbit ID", "value": "", "error": "Orbit ID is required for new schedule creation."})

                    support_type = row_dict.get("support_type")
                    if not support_type:
                        row_errors.append({"field": "Support Type", "value": "", "error": "Support Type is required."})

                    country = row_dict.get("country")
                    if not country:
                        row_errors.append({"field": "Country", "value": "", "error": "Country is required."})

                    if not start_date:
                        row_errors.append({"field": "Start Date", "value": "", "error": "Start Date is required."})

                    if row_errors:
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    eng_info = orbit_to_engineer.get(norm_str(orbit_id))
                    if not eng_info:
                        row_errors.append({"field": "Orbit ID", "value": str(orbit_id), "error": f"Engineer with Orbit ID '{orbit_id}' does not exist in the selected company."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    engineer_id, resolved_engineer_name = eng_info
                    row_dict["engineer_id"] = engineer_id
                    row_dict["resolved_engineer_name"] = resolved_engineer_name
                    row_dict["schedule_status"] = row_dict.get("schedule_status") or "Upcoming"

                    fab_city = row_dict.get("fab_city")
                    fab_site = row_dict.get("fab_site")

                    row_key = (norm_uuid(engineer_id), norm_str(support_type), norm_str(country), norm_str(fab_city), norm_str(fab_site), norm_date(start_date), norm_date(end_date))
                    if row_key in seen_keys:
                        row_dict["duplicate_key"] = f"EngineerID: {engineer_id}, Type: {support_type}, Location: {country}/{fab_city}/{fab_site}"
                        duplicates_list.append(row_dict)
                        continue
                    seen_keys.add(row_key)

                    valid_rows_to_insert.append(row_dict)

            # Persist Inserts & Updates
            imported_count = 0
            failed_count = 0
            try:
                schedules_to_add = []
                for item in valid_rows_to_insert:
                    db_sched = Schedule(
                        schedule_id=uuid_pkg.uuid4(),
                        engineer_id=item["engineer_id"],
                        owner_id=item.get("owner_id"),
                        support_type=item["support_type"],
                        country=item["country"],
                        fab_city=item.get("fab_city"),
                        fab_site=item.get("fab_site"),
                        start_date=item["start_date"],
                        end_date=item.get("end_date"),
                        schedule_status=item.get("schedule_status") or "Upcoming",
                        remarks=item.get("remarks"),
                        comment_status=item.get("comment_status") or "UNADDRESSED",
                        created_at=datetime.utcnow(),
                        updated_at=datetime.utcnow()
                    )
                    schedules_to_add.append(db_sched)
                if schedules_to_add:
                    db.add_all(schedules_to_add)
                db.commit()
                imported_count = len(valid_rows_to_insert) + len(existing_list)
            except Exception as insert_err:
                db.rollback()
                failed_count = len(valid_rows_to_insert) + len(existing_list)
                bulk_upload_service.update_bulk_upload(
                    db,
                    upload_id=upload_id,
                    status="FAILED",
                    failed_rows=failed_count
                )
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail=f"Database ingestion failed: {str(insert_err)}"
                )

            # Generate Report
            report_wb = openpyxl.Workbook()
            ws_summary = report_wb.active
            ws_summary.title = "Summary"
            ws_summary.append(["ORMP Schedule Bulk Ingestion Report"])
            ws_summary.append([])
            ws_summary.append(["File Name", file.filename])
            ws_summary.append(["Uploaded By", current_user.full_name])
            ws_summary.append(["Upload Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            ws_summary.append(["Target Company", company.company_name])
            ws_summary.append(["Company UUID", str(company.company_id)])
            ws_summary.append([])
            ws_summary.append(["Metric", "Count"])
            ws_summary.append(["Total Rows", total_rows])
            ws_summary.append(["Inserted Records", len(valid_rows_to_insert)])
            ws_summary.append(["Updated Records", len(existing_list)])
            ws_summary.append(["Unchanged Records", len(unchanged_list)])
            ws_summary.append(["Error Rows", len(errors_list)])
            ws_summary.append(["Duplicate Rows", len(duplicates_list)])
            ws_summary.append(["Warning Rows", 0])

            for col in ws_summary.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws_summary.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)

            headers_valid = [
                "Excel Row", "Schedule ID", "Orbit ID", "Engineer Name", "Support Type", "Country", "Fab City", 
                "Fab Site", "Start Date", "End Date", "Schedule Status", "Remarks", "Status"
            ]

            ws_valid = report_wb.create_sheet(title="Valid Records")
            ws_valid.append(headers_valid)
            for r in valid_rows_to_insert:
                ws_valid.append([
                    r["excel_row"],
                    str(r.get("schedule_id") or ""),
                    r.get("orbit_id"),
                    r.get("resolved_engineer_name"),
                    r.get("support_type"),
                    r.get("country"),
                    r.get("fab_city"),
                    r.get("fab_site"),
                    str(r.get("start_date")) if r.get("start_date") else "",
                    str(r.get("end_date")) if r.get("end_date") else "",
                    r.get("schedule_status"),
                    r.get("remarks"),
                    "INSERTED"
                ])

            ws_updated = report_wb.create_sheet(title="Updated Records")
            ws_updated.append(["Excel Row", "Schedule ID", "Orbit ID", "Engineer Name", "Action Status", "Changed Columns", "Support Type", "Country", "Fab City", "Fab Site", "Status", "Remarks"])
            for r in existing_list:
                ws_updated.append([
                    r["excel_row"],
                    str(r.get("schedule_id") or ""),
                    r.get("orbit_id") or "",
                    r.get("resolved_engineer_name") or "",
                    "UPDATED",
                    r.get("changed_fields") or "",
                    r.get("support_type"),
                    r.get("country"),
                    r.get("fab_city"),
                    r.get("fab_site"),
                    r.get("schedule_status"),
                    r.get("remarks")
                ])

            ws_unchanged = report_wb.create_sheet(title="Unchanged Records")
            ws_unchanged.append(["Excel Row", "Schedule ID", "Orbit ID", "Engineer Name", "Action Status", "Details", "Support Type", "Country", "Fab City", "Fab Site", "Status", "Remarks"])
            for r in unchanged_list:
                ws_unchanged.append([
                    r["excel_row"],
                    str(r.get("schedule_id") or ""),
                    r.get("orbit_id") or "",
                    r.get("resolved_engineer_name") or "",
                    "UNCHANGED",
                    "All supplied values match database",
                    r.get("support_type"),
                    r.get("country"),
                    r.get("fab_city"),
                    r.get("fab_site"),
                    r.get("schedule_status"),
                    r.get("remarks")
                ])

            ws_errors = report_wb.create_sheet(title="Errors")
            ws_errors.append(["Excel Row", "Orbit ID", "Field", "Value", "Error"])
            for r in errors_list:
                for err in r.get("errors", []):
                    ws_errors.append([
                        r["excel_row"],
                        r.get("orbit_id") or "",
                        err.get("field") or "",
                        err.get("value") or "",
                        err.get("error") or ""
                    ])

            ws_dups = report_wb.create_sheet(title="Duplicates")
            ws_dups.append(["Excel Row", "Orbit ID", "Duplicate Key", "Duplicate Rows", "Reason"])
            for r in duplicates_list:
                ws_dups.append([
                    r["excel_row"],
                    r.get("orbit_id") or "",
                    r.get("duplicate_key") or "",
                    "",
                    "Duplicate Schedule row in Excel sheet"
                ])

            ws_warn = report_wb.create_sheet(title="Warnings")
            ws_warn.append(["Excel Row", "Orbit ID", "Field", "Value", "Warning"])

            for sheet_obj in (ws_valid, ws_updated, ws_unchanged, ws_errors, ws_dups, ws_warn):
                for col in sheet_obj.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    sheet_obj.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)

            os.makedirs(TEMP_REPORTS_DIR, exist_ok=True)
            report_filename = f"validation_report_{uuid_pkg.uuid4()}.xlsx"
            report_path = os.path.join(TEMP_REPORTS_DIR, report_filename)
            report_wb.save(report_path)

            final_status = "COMPLETED"
            if len(errors_list) > 0 or len(duplicates_list) > 0:
                final_status = "COMPLETED_WITH_ERRORS"

            bulk_upload_service.update_bulk_upload(
                db,
                upload_id=upload_id,
                status=final_status,
                report_file=report_filename,
                imported_rows=imported_count,
                failed_rows=failed_count
            )

            ingested_msg = f"Ingested {len(valid_rows_to_insert)} new schedule records. Updated {len(existing_list)} existing schedule records. {len(unchanged_list)} records were unchanged."
            if errors_list or duplicates_list:
                ingested_msg += " Some rows had validation errors or duplicates. See the validation report for details."

            return {
                "success": True,
                "rowsProcessed": total_rows,
                "errorsCount": len(errors_list),
                "inserted": len(valid_rows_to_insert),
                "updated": len(existing_list),
                "unchanged": len(unchanged_list),
                "message": ingested_msg,
                "report_url": f"/api/upload/download-report/{report_filename}"
            }

        # =========================================================================
        # MODULE 3: UP-VISA
        # =========================================================================
        if module_id == "up-visa":
            start_time = time.perf_counter()
            try:
                contents = await file.read()
                wb = openpyxl.load_workbook(io.BytesIO(contents))
            except Exception:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Failed to parse Excel file. Please ensure it is a valid .xlsx file."
                )

            visa_sheet_name = None
            for name in wb.sheetnames:
                norm_name = name.strip().lower()
                if norm_name in ("visa", "visas", "visa details", "visa matrix"):
                    visa_sheet_name = name
                    break

            if not visa_sheet_name:
                if len(wb.sheetnames) == 1:
                    visa_sheet_name = wb.sheetnames[0]
                else:
                    raise HTTPException(
                        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                        detail="Excel workbook must contain a 'Visa' or 'Visas' sheet."
                    )

            sheet = wb[visa_sheet_name]
            first_row = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
            col_indices = {}
            for idx, val in enumerate(first_row):
                if val is not None:
                    norm = normalize_header(val)
                    mapped_field = VISA_HEADER_MAP.get(norm)
                    if mapped_field:
                        col_indices[mapped_field] = idx + 1

            raw_rows = []
            for r in range(2, sheet.max_row + 1):
                is_blank = True
                for c in range(1, sheet.max_column + 1):
                    val = sheet.cell(row=r, column=c).value
                    if val is not None and str(val).strip() != "":
                        is_blank = False
                        break
                if is_blank:
                    continue

                row_dict = {"excel_row": r}
                original_engineer_name = None
                for idx, val in enumerate(first_row):
                    if val is not None:
                        norm = normalize_header(val)
                        if norm in ("engineername", "name"):
                            original_engineer_name = clean_val(sheet.cell(row=r, column=idx + 1).value)
                            break
                row_dict["original_engineer_name"] = original_engineer_name

                for field, col_idx in col_indices.items():
                    row_dict[field] = clean_val(sheet.cell(row=r, column=col_idx).value)

                raw_rows.append(row_dict)

            if not raw_rows:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="The Visa sheet is empty or contains no rows."
                )

            unique_orbit_ids = {
                norm_str(row.get("orbit_id"))
                for row in raw_rows
                if row.get("orbit_id") and norm_str(row.get("orbit_id")) != ""
            }

            db_engineers = []
            if unique_orbit_ids:
                db_engineers = db.scalars(
                    select(Engineer).where(
                        func.lower(Engineer.orbit_id).in_(list(unique_orbit_ids)),
                        Engineer.company_id == target_company_id
                    )
                ).all()

            orbit_to_engineer = {
                norm_str(eng.orbit_id): (eng.engineer_id, eng.engineer_name)
                for eng in db_engineers
            }

            errors_list = []
            duplicates_list = []
            existing_list = []
            unchanged_list = []
            valid_rows_to_insert = []
            seen_keys = set()

            total_rows = len(raw_rows)

            for row_dict in raw_rows:
                row_errors = []
                raw_pk = row_dict.get("visa_id")
                parsed_pk = None

                if raw_pk is not None and str(raw_pk).strip() != "":
                    try:
                        parsed_pk = parse_uuid_safe(raw_pk)
                    except ValueError as ve:
                        row_errors.append({"field": "Visa ID", "value": str(raw_pk), "error": str(ve)})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                applied_on = None
                if row_dict.get("applied_on") is not None:
                    try:
                        applied_on = parse_date(row_dict["applied_on"])
                    except ValueError:
                        row_errors.append({"field": "Applied On", "value": str(row_dict["applied_on"]), "error": "Invalid applied date format."})

                visa_start_date = None
                if row_dict.get("visa_start_date") is not None:
                    try:
                        visa_start_date = parse_date(row_dict["visa_start_date"])
                    except ValueError:
                        row_errors.append({"field": "Start Date", "value": str(row_dict["visa_start_date"]), "error": "Invalid start date format."})

                visa_end_date = None
                if row_dict.get("visa_end_date") is not None:
                    try:
                        visa_end_date = parse_date(row_dict["visa_end_date"])
                    except ValueError:
                        row_errors.append({"field": "End Date / Expiry Date", "value": str(row_dict["visa_end_date"]), "error": "Invalid end date format."})

                if visa_start_date and visa_end_date and visa_end_date < visa_start_date:
                    row_errors.append({"field": "End Date / Expiry Date", "value": str(row_dict["visa_end_date"]), "error": "visa_end_date should not be earlier than visa_start_date"})

                resolved_owner_id = resolve_owner(row_dict.get("owner"))
                row_dict["owner_id"] = resolved_owner_id
                row_dict["applied_on"] = applied_on
                row_dict["visa_start_date"] = visa_start_date
                row_dict["visa_end_date"] = visa_end_date

                if row_errors:
                    row_dict["errors"] = row_errors
                    errors_list.append(row_dict)
                    continue

                # PRIMARY KEY UPDATE PATH
                if parsed_pk is not None:
                    db_v = db.get(Visa, parsed_pk)
                    if not db_v:
                        row_errors.append({"field": "Visa ID", "value": str(parsed_pk), "error": f"Record with Visa ID '{parsed_pk}' was not found. Cannot update non-existent record ID."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    eng = db.get(Engineer, db_v.engineer_id)
                    if not eng or eng.company_id != target_company_id:
                        row_errors.append({"field": "Visa ID", "value": str(parsed_pk), "error": f"Record with Visa ID '{parsed_pk}' belongs to another company."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    row_dict["resolved_engineer_name"] = eng.engineer_name
                    row_dict["orbit_id"] = eng.orbit_id

                    changes = []
                    field_specs = [
                        ("country", "Country", db_v.country),
                        ("visa_type", "Visa Type", db_v.visa_type),
                        ("applied_on", "Applied On", db_v.applied_on),
                        ("visa_start_date", "Start Date", db_v.visa_start_date),
                        ("visa_end_date", "End Date", db_v.visa_end_date),
                        ("comments", "Comments", db_v.comments),
                        ("comment_status", "Comment Status", db_v.comment_status),
                        ("owner_id", "Owner ID", db_v.owner_id)
                    ]

                    for f_key, f_label, cur_val in field_specs:
                        if f_key in col_indices or (f_key == "owner_id" and "owner" in col_indices):
                            new_val = row_dict.get(f_key)
                            if not values_are_equal(cur_val, new_val):
                                changes.append(f"{f_label}: '{cur_val}' -> '{new_val}'")
                                setattr(db_v, f_key, new_val)

                    if changes:
                        db_v.updated_at = datetime.utcnow()
                        row_dict["update_status"] = "UPDATED"
                        row_dict["changed_fields"] = "; ".join(changes)
                        existing_list.append(row_dict)
                    else:
                        row_dict["update_status"] = "UNCHANGED"
                        row_dict["changed_fields"] = "No fields modified"
                        unchanged_list.append(row_dict)

                # NEW RECORD INSERTION PATH
                else:
                    orbit_id = row_dict.get("orbit_id")
                    if not orbit_id:
                        row_errors.append({"field": "Orbit ID", "value": "", "error": "Orbit ID is required for new visa creation."})

                    country = row_dict.get("country")
                    if not country:
                        row_errors.append({"field": "Country", "value": "", "error": "Country is required."})

                    if row_errors:
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    eng_info = orbit_to_engineer.get(norm_str(orbit_id))
                    if not eng_info:
                        row_errors.append({"field": "Orbit ID", "value": str(orbit_id), "error": f"Engineer with Orbit ID '{orbit_id}' does not exist in the selected company."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    engineer_id, resolved_engineer_name = eng_info
                    row_dict["engineer_id"] = engineer_id
                    row_dict["resolved_engineer_name"] = resolved_engineer_name

                    visa_type = row_dict.get("visa_type") or ""
                    row_key = (norm_uuid(engineer_id), (country or "").strip().lower(), (visa_type or "").strip().lower())
                    if row_key in seen_keys:
                        row_dict["duplicate_key"] = f"EngineerID: {engineer_id}, Country: {country}, Type: {visa_type}"
                        duplicates_list.append(row_dict)
                        continue
                    seen_keys.add(row_key)

                    valid_rows_to_insert.append(row_dict)

            # Persist Inserts & Updates
            imported_count = 0
            failed_count = 0
            try:
                visas_to_add = []
                for item in valid_rows_to_insert:
                    db_v = Visa(
                        visa_id=uuid_pkg.uuid4(),
                        engineer_id=item["engineer_id"],
                        owner_id=item.get("owner_id"),
                        country=item["country"],
                        visa_type=item.get("visa_type"),
                        applied_on=item.get("applied_on"),
                        visa_start_date=item.get("visa_start_date"),
                        visa_end_date=item.get("visa_end_date"),
                        comments=item.get("comments"),
                        comment_status=item.get("comment_status") or "UNADDRESSED",
                        created_at=datetime.utcnow(),
                        updated_at=datetime.utcnow()
                    )
                    visas_to_add.append(db_v)
                if visas_to_add:
                    db.add_all(visas_to_add)

                db.commit()
                imported_count = len(valid_rows_to_insert) + len(existing_list)
            except Exception as insert_err:
                db.rollback()
                failed_count = len(valid_rows_to_insert) + len(existing_list)
                bulk_upload_service.update_bulk_upload(
                    db,
                    upload_id=upload_id,
                    status="FAILED",
                    failed_rows=failed_count
                )
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail=f"Database ingestion failed: {str(insert_err)}"
                )

            # Generate Report
            report_wb = openpyxl.Workbook()
            ws_summary = report_wb.active
            ws_summary.title = "Summary"
            ws_summary.append(["ORMP Visa Bulk Ingestion Report"])
            ws_summary.append([])
            ws_summary.append(["File Name", file.filename])
            ws_summary.append(["Uploaded By", current_user.full_name])
            ws_summary.append(["Upload Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            ws_summary.append(["Target Company", company.company_name])
            ws_summary.append(["Company UUID", str(company.company_id)])
            ws_summary.append([])
            ws_summary.append(["Metric", "Count"])
            ws_summary.append(["Total Rows", total_rows])
            ws_summary.append(["Inserted Records", len(valid_rows_to_insert)])
            ws_summary.append(["Updated Records", len(existing_list)])
            ws_summary.append(["Unchanged Records", len(unchanged_list)])
            ws_summary.append(["Error Rows", len(errors_list)])
            ws_summary.append(["Duplicate Rows", len(duplicates_list)])
            ws_summary.append(["Warning Rows", 0])

            for col in ws_summary.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws_summary.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)

            headers_valid = [
                "Excel Row", "Visa ID", "Orbit ID", "Engineer Name", "Country", "Visa Type", 
                "Applied On", "Start Date", "End Date", "Comments", "Status"
            ]

            ws_valid = report_wb.create_sheet(title="Valid Records")
            ws_valid.append(headers_valid)
            for r in valid_rows_to_insert:
                ws_valid.append([
                    r["excel_row"],
                    str(r.get("visa_id") or ""),
                    r.get("orbit_id"),
                    r.get("resolved_engineer_name"),
                    r.get("country"),
                    r.get("visa_type"),
                    str(r.get("applied_on")) if r.get("applied_on") else "",
                    str(r.get("visa_start_date")) if r.get("visa_start_date") else "",
                    str(r.get("visa_end_date")) if r.get("visa_end_date") else "",
                    r.get("comments"),
                    "INSERTED"
                ])

            ws_updated = report_wb.create_sheet(title="Updated Records")
            ws_updated.append(["Excel Row", "Visa ID", "Orbit ID", "Engineer Name", "Action Status", "Changed Columns", "Country", "Visa Type", "Comments"])
            for r in existing_list:
                ws_updated.append([
                    r["excel_row"],
                    str(r.get("visa_id") or ""),
                    r.get("orbit_id") or "",
                    r.get("resolved_engineer_name") or r.get("original_engineer_name") or "",
                    "UPDATED",
                    r.get("changed_fields") or "",
                    r.get("country"),
                    r.get("visa_type"),
                    r.get("comments")
                ])

            ws_unchanged = report_wb.create_sheet(title="Unchanged Records")
            ws_unchanged.append(["Excel Row", "Visa ID", "Orbit ID", "Engineer Name", "Action Status", "Details", "Country", "Visa Type", "Comments"])
            for r in unchanged_list:
                ws_unchanged.append([
                    r["excel_row"],
                    str(r.get("visa_id") or ""),
                    r.get("orbit_id") or "",
                    r.get("resolved_engineer_name") or r.get("original_engineer_name") or "",
                    "UNCHANGED",
                    "All supplied values match database",
                    r.get("country"),
                    r.get("visa_type"),
                    r.get("comments")
                ])

            ws_errors = report_wb.create_sheet(title="Errors")
            ws_errors.append(["Excel Row", "Orbit ID", "Field", "Value", "Error"])
            for r in errors_list:
                for err in r.get("errors", []):
                    ws_errors.append([
                        r["excel_row"],
                        r.get("orbit_id") or "",
                        err.get("field") or "",
                        err.get("value") or "",
                        err.get("error") or ""
                    ])

            ws_dups = report_wb.create_sheet(title="Duplicates")
            ws_dups.append(["Excel Row", "Orbit ID", "Duplicate Key", "Reason"])
            for r in duplicates_list:
                ws_dups.append([
                    r["excel_row"],
                    r.get("orbit_id") or "",
                    r.get("duplicate_key") or "",
                    "Duplicate Visa row within Excel sheet"
                ])

            ws_warn = report_wb.create_sheet(title="Warnings")
            ws_warn.append(["Excel Row", "Orbit ID", "Field", "Value", "Warning"])

            for sheet_obj in (ws_valid, ws_updated, ws_unchanged, ws_errors, ws_dups, ws_warn):
                for col in sheet_obj.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    sheet_obj.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)

            os.makedirs(TEMP_REPORTS_DIR, exist_ok=True)
            report_filename = f"validation_report_{uuid_pkg.uuid4()}.xlsx"
            report_path = os.path.join(TEMP_REPORTS_DIR, report_filename)
            report_wb.save(report_path)

            final_status = "COMPLETED"
            if len(errors_list) > 0 or len(duplicates_list) > 0:
                final_status = "COMPLETED_WITH_ERRORS"

            bulk_upload_service.update_bulk_upload(
                db,
                upload_id=upload_id,
                status=final_status,
                report_file=report_filename,
                imported_rows=imported_count,
                failed_rows=failed_count
            )

            ingested_msg = f"Ingested {len(valid_rows_to_insert)} new visa records. Updated {len(existing_list)} existing visa records. {len(unchanged_list)} records were unchanged."
            if errors_list or duplicates_list:
                ingested_msg += " Some rows had validation errors or duplicates. See the validation report for details."

            return {
                "success": True,
                "rowsProcessed": total_rows,
                "errorsCount": len(errors_list),
                "inserted": len(valid_rows_to_insert),
                "updated": len(existing_list),
                "unchanged": len(unchanged_list),
                "message": ingested_msg,
                "report_url": f"/api/upload/download-report/{report_filename}"
            }

        # =========================================================================
        # MODULE 4: UP-TRAVEL
        # =========================================================================
        if module_id == "up-travel":
            start_time = time.perf_counter()
            try:
                contents = await file.read()
                wb = openpyxl.load_workbook(io.BytesIO(contents))
            except Exception:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Failed to parse Excel file. Please ensure it is a valid .xlsx file."
                )

            travel_sheet_name = None
            for name in wb.sheetnames:
                norm_name = name.strip().lower()
                if norm_name in ("travel", "travels", "travel details", "travel arrangements", "travel itinerary"):
                    travel_sheet_name = name
                    break

            if not travel_sheet_name:
                if len(wb.sheetnames) == 1:
                    travel_sheet_name = wb.sheetnames[0]
                else:
                    raise HTTPException(
                        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                        detail="Excel workbook must contain a 'Travel' or 'Travel Details' sheet."
                    )

            sheet = wb[travel_sheet_name]
            first_row = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
            col_indices = {}
            for idx, val in enumerate(first_row):
                if val is not None:
                    norm = normalize_header(val)
                    mapped_field = TRAVEL_HEADER_MAP.get(norm)
                    if mapped_field:
                        col_indices[mapped_field] = idx + 1

            raw_rows = []
            for r in range(2, sheet.max_row + 1):
                is_blank = True
                for c in range(1, sheet.max_column + 1):
                    val = sheet.cell(row=r, column=c).value
                    if val is not None and str(val).strip() != "":
                        is_blank = False
                        break
                if is_blank:
                    continue

                row_dict = {"excel_row": r}
                original_engineer_name = None
                for idx, val in enumerate(first_row):
                    if val is not None:
                        norm = normalize_header(val)
                        if norm in ("engineername", "name"):
                            original_engineer_name = clean_val(sheet.cell(row=r, column=idx + 1).value)
                            break
                row_dict["original_engineer_name"] = original_engineer_name

                for field, col_idx in col_indices.items():
                    row_dict[field] = clean_val(sheet.cell(row=r, column=col_idx).value)

                raw_rows.append(row_dict)

            if not raw_rows:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="The Travel sheet is empty or contains no rows."
                )

            unique_orbit_ids = {
                norm_str(row.get("orbit_id"))
                for row in raw_rows
                if row.get("orbit_id") and norm_str(row.get("orbit_id")) != ""
            }

            db_engineers = []
            if unique_orbit_ids:
                db_engineers = db.scalars(
                    select(Engineer).where(
                        func.lower(Engineer.orbit_id).in_(list(unique_orbit_ids)),
                        Engineer.company_id == target_company_id
                    )
                ).all()

            orbit_to_engineer = {
                norm_str(eng.orbit_id): (eng.engineer_id, eng.engineer_name)
                for eng in db_engineers
            }

            resolved_engineer_ids = {val[0] for val in orbit_to_engineer.values()}
            db_schedules = []
            if resolved_engineer_ids:
                db_schedules = db.scalars(
                    select(Schedule).where(Schedule.engineer_id.in_(list(resolved_engineer_ids)))
                ).all()

            engineer_schedules = {}
            for sch in db_schedules:
                engineer_schedules.setdefault(sch.engineer_id, []).append(sch)

            errors_list = []
            duplicates_list = []
            existing_list = []
            unchanged_list = []
            valid_rows_to_insert = []
            new_schedules_created = []
            seen_keys = set()

            total_rows = len(raw_rows)

            for row_dict in raw_rows:
                row_errors = []
                raw_pk = row_dict.get("travel_id")
                parsed_pk = None

                if raw_pk is not None and str(raw_pk).strip() != "":
                    try:
                        parsed_pk = parse_uuid_safe(raw_pk)
                    except ValueError as ve:
                        row_errors.append({"field": "Travel ID", "value": str(raw_pk), "error": str(ve)})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                booking_date = None
                if row_dict.get("booking_date") is not None:
                    try:
                        booking_date = parse_date(row_dict["booking_date"])
                    except ValueError:
                        row_errors.append({"field": "Booking Date", "value": str(row_dict["booking_date"]), "error": "Invalid booking date format."})

                travel_date = None
                if row_dict.get("travel_date") is not None:
                    try:
                        travel_date = parse_date(row_dict["travel_date"])
                    except ValueError:
                        row_errors.append({"field": "Travel Date", "value": str(row_dict["travel_date"]), "error": "Invalid travel date format."})

                if booking_date and travel_date and travel_date < booking_date:
                    row_errors.append({"field": "Travel Date", "value": str(row_dict["travel_date"]), "error": "travel_date should not be earlier than booking_date"})

                resolved_owner_id = resolve_owner(row_dict.get("owner"))
                row_dict["owner_id"] = resolved_owner_id
                row_dict["booking_date"] = booking_date
                row_dict["travel_date"] = travel_date
                purpose = row_dict.get("purpose") or "Customer Support"
                row_dict["purpose"] = purpose

                if row_errors:
                    row_dict["errors"] = row_errors
                    errors_list.append(row_dict)
                    continue

                # PRIMARY KEY UPDATE PATH
                if parsed_pk is not None:
                    db_tr = db.get(Travel, parsed_pk)
                    if not db_tr:
                        row_errors.append({"field": "Travel ID", "value": str(parsed_pk), "error": f"Record with Travel ID '{parsed_pk}' was not found. Cannot update non-existent record ID."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    sch = db.get(Schedule, db_tr.schedule_id)
                    eng = db.get(Engineer, sch.engineer_id) if sch else None
                    if not eng or eng.company_id != target_company_id:
                        row_errors.append({"field": "Travel ID", "value": str(parsed_pk), "error": f"Record with Travel ID '{parsed_pk}' belongs to another company."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    row_dict["resolved_engineer_name"] = eng.engineer_name
                    row_dict["orbit_id"] = eng.orbit_id

                    changes = []
                    field_specs = [
                        ("booking_date", "Booking Date", db_tr.booking_date),
                        ("travel_date", "Travel Date", db_tr.travel_date),
                        ("purpose", "Purpose", db_tr.purpose),
                        ("comments", "Comments", db_tr.comments),
                        ("owner_id", "Owner ID", db_tr.owner_id)
                    ]

                    for f_key, f_label, cur_val in field_specs:
                        if f_key in col_indices or (f_key == "owner_id" and "owner" in col_indices):
                            new_val = row_dict.get(f_key)
                            if not values_are_equal(cur_val, new_val):
                                changes.append(f"{f_label}: '{cur_val}' -> '{new_val}'")
                                setattr(db_tr, f_key, new_val)

                    if changes:
                        db_tr.updated_at = datetime.utcnow()
                        row_dict["update_status"] = "UPDATED"
                        row_dict["changed_fields"] = "; ".join(changes)
                        existing_list.append(row_dict)
                    else:
                        row_dict["update_status"] = "UNCHANGED"
                        row_dict["changed_fields"] = "No fields modified"
                        unchanged_list.append(row_dict)

                # NEW RECORD INSERTION PATH
                else:
                    orbit_id = row_dict.get("orbit_id")
                    if not orbit_id:
                        row_errors.append({"field": "Orbit ID", "value": "", "error": "Orbit ID is required for new travel arrangement."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    eng_info = orbit_to_engineer.get(norm_str(orbit_id))
                    if not eng_info:
                        row_errors.append({"field": "Orbit ID", "value": str(orbit_id), "error": f"Engineer with Orbit ID '{orbit_id}' does not exist in the selected company."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    engineer_id, resolved_engineer_name = eng_info
                    row_dict["engineer_id"] = engineer_id
                    row_dict["resolved_engineer_name"] = resolved_engineer_name

                    schedules_for_eng = engineer_schedules.get(engineer_id, [])
                    if schedules_for_eng:
                        target_schedule = schedules_for_eng[0]
                    else:
                        target_schedule = Schedule(
                            schedule_id=uuid_pkg.uuid4(),
                            engineer_id=engineer_id,
                            support_type=row_dict.get("support_type") or "Customer Support",
                            country=row_dict.get("country") or "Global",
                            fab_city=row_dict.get("fab_city"),
                            fab_site=row_dict.get("fab_site"),
                            start_date=travel_date or date.today(),
                            schedule_status="Upcoming",
                            created_at=datetime.utcnow(),
                            updated_at=datetime.utcnow()
                        )
                        engineer_schedules.setdefault(engineer_id, []).append(target_schedule)
                        new_schedules_created.append(target_schedule)

                    schedule_id = target_schedule.schedule_id
                    row_dict["schedule_id"] = schedule_id

                    row_key = (norm_uuid(schedule_id), norm_date(travel_date), (purpose or "").strip().lower())
                    if row_key in seen_keys:
                        row_dict["duplicate_key"] = f"OrbitID: {orbit_id}, TravelDate: {travel_date}, Purpose: {purpose}"
                        duplicates_list.append(row_dict)
                        continue
                    seen_keys.add(row_key)

                    valid_rows_to_insert.append(row_dict)

            # Persist Inserts & Updates
            imported_count = 0
            failed_count = 0
            try:
                if new_schedules_created:
                    db.add_all(new_schedules_created)

                travels_to_add = []
                for item in valid_rows_to_insert:
                    db_tr = Travel(
                        travel_id=uuid_pkg.uuid4(),
                        schedule_id=item["schedule_id"],
                        owner_id=item.get("owner_id"),
                        booking_date=item.get("booking_date"),
                        travel_date=item.get("travel_date"),
                        purpose=item.get("purpose"),
                        comments=item.get("comments"),
                        created_at=datetime.utcnow(),
                        updated_at=datetime.utcnow()
                    )
                    travels_to_add.append(db_tr)
                if travels_to_add:
                    db.add_all(travels_to_add)

                db.commit()
                imported_count = len(valid_rows_to_insert) + len(existing_list)
            except Exception as insert_err:
                db.rollback()
                failed_count = len(valid_rows_to_insert) + len(existing_list)
                bulk_upload_service.update_bulk_upload(
                    db,
                    upload_id=upload_id,
                    status="FAILED",
                    failed_rows=failed_count
                )
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail=f"Database ingestion failed: {str(insert_err)}"
                )

            # Generate Report
            report_wb = openpyxl.Workbook()
            ws_summary = report_wb.active
            ws_summary.title = "Summary"
            ws_summary.append(["ORMP Travel Bulk Ingestion Report"])
            ws_summary.append([])
            ws_summary.append(["File Name", file.filename])
            ws_summary.append(["Uploaded By", current_user.full_name])
            ws_summary.append(["Upload Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            ws_summary.append(["Target Company", company.company_name])
            ws_summary.append(["Company UUID", str(company.company_id)])
            ws_summary.append([])
            ws_summary.append(["Metric", "Count"])
            ws_summary.append(["Total Rows", total_rows])
            ws_summary.append(["Inserted Records", len(valid_rows_to_insert)])
            ws_summary.append(["Updated Records", len(existing_list)])
            ws_summary.append(["Unchanged Records", len(unchanged_list)])
            ws_summary.append(["Error Rows", len(errors_list)])
            ws_summary.append(["Duplicate Rows", len(duplicates_list)])
            ws_summary.append(["Warning Rows", 0])

            for col in ws_summary.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws_summary.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)

            headers_valid = [
                "Excel Row", "Travel ID", "Orbit ID", "Engineer Name", "Booking Date", "Travel Date", 
                "Purpose", "Comments", "Status"
            ]

            ws_valid = report_wb.create_sheet(title="Valid Records")
            ws_valid.append(headers_valid)
            for r in valid_rows_to_insert:
                ws_valid.append([
                    r["excel_row"],
                    str(r.get("travel_id") or ""),
                    r.get("orbit_id"),
                    r.get("resolved_engineer_name"),
                    str(r.get("booking_date")) if r.get("booking_date") else "",
                    str(r.get("travel_date")) if r.get("travel_date") else "",
                    r.get("purpose"),
                    r.get("comments"),
                    "INSERTED"
                ])

            ws_updated = report_wb.create_sheet(title="Updated Records")
            ws_updated.append(["Excel Row", "Travel ID", "Orbit ID", "Engineer Name", "Action Status", "Changed Columns", "Booking Date", "Travel Date", "Purpose", "Comments"])
            for r in existing_list:
                ws_updated.append([
                    r["excel_row"],
                    str(r.get("travel_id") or ""),
                    r.get("orbit_id") or "",
                    r.get("resolved_engineer_name") or r.get("original_engineer_name") or "",
                    "UPDATED",
                    r.get("changed_fields") or "",
                    str(r.get("booking_date")) if r.get("booking_date") else "",
                    str(r.get("travel_date")) if r.get("travel_date") else "",
                    r.get("purpose"),
                    r.get("comments")
                ])

            ws_unchanged = report_wb.create_sheet(title="Unchanged Records")
            ws_unchanged.append(["Excel Row", "Travel ID", "Orbit ID", "Engineer Name", "Action Status", "Details", "Booking Date", "Travel Date", "Purpose", "Comments"])
            for r in unchanged_list:
                ws_unchanged.append([
                    r["excel_row"],
                    str(r.get("travel_id") or ""),
                    r.get("orbit_id") or "",
                    r.get("resolved_engineer_name") or r.get("original_engineer_name") or "",
                    "UNCHANGED",
                    "All supplied values match database",
                    str(r.get("booking_date")) if r.get("booking_date") else "",
                    str(r.get("travel_date")) if r.get("travel_date") else "",
                    r.get("purpose"),
                    r.get("comments")
                ])

            ws_errors = report_wb.create_sheet(title="Errors")
            ws_errors.append(["Excel Row", "Orbit ID", "Field", "Value", "Error"])
            for r in errors_list:
                for err in r.get("errors", []):
                    ws_errors.append([
                        r["excel_row"],
                        r.get("orbit_id") or "",
                        err.get("field") or "",
                        err.get("value") or "",
                        err.get("error") or ""
                    ])

            ws_dups = report_wb.create_sheet(title="Duplicates")
            ws_dups.append(["Excel Row", "Orbit ID", "Duplicate Key", "Reason"])
            for r in duplicates_list:
                ws_dups.append([
                    r["excel_row"],
                    r.get("orbit_id") or "",
                    r.get("duplicate_key") or "",
                    "Duplicate Travel row within Excel sheet"
                ])

            ws_warn = report_wb.create_sheet(title="Warnings")
            ws_warn.append(["Excel Row", "Orbit ID", "Field", "Value", "Warning"])

            for sheet_obj in (ws_valid, ws_updated, ws_unchanged, ws_errors, ws_dups, ws_warn):
                for col in sheet_obj.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    sheet_obj.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)

            os.makedirs(TEMP_REPORTS_DIR, exist_ok=True)
            report_filename = f"validation_report_{uuid_pkg.uuid4()}.xlsx"
            report_path = os.path.join(TEMP_REPORTS_DIR, report_filename)
            report_wb.save(report_path)

            final_status = "COMPLETED"
            if len(errors_list) > 0 or len(duplicates_list) > 0:
                final_status = "COMPLETED_WITH_ERRORS"

            bulk_upload_service.update_bulk_upload(
                db,
                upload_id=upload_id,
                status=final_status,
                report_file=report_filename,
                imported_rows=imported_count,
                failed_rows=failed_count
            )

            ingested_msg = f"Ingested {len(valid_rows_to_insert)} new travel records. Updated {len(existing_list)} existing travel records. {len(unchanged_list)} records were unchanged."
            if errors_list or duplicates_list:
                ingested_msg += " Some rows had validation errors or duplicates. See the validation report for details."

            return {
                "success": True,
                "rowsProcessed": total_rows,
                "errorsCount": len(errors_list),
                "inserted": len(valid_rows_to_insert),
                "updated": len(existing_list),
                "unchanged": len(unchanged_list),
                "message": ingested_msg,
                "report_url": f"/api/upload/download-report/{report_filename}"
            }

        # =========================================================================
        # MODULE 5: UP-PERFORMANCE
        # =========================================================================
        if module_id == "up-performance":
            start_time = time.perf_counter()
            try:
                contents = await file.read()
                wb = openpyxl.load_workbook(io.BytesIO(contents))
            except Exception:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Failed to parse Excel file. Please ensure it is a valid .xlsx file."
                )

            perf_sheet_name = None
            for name in wb.sheetnames:
                norm_name = name.strip().lower()
                if norm_name in ("performance", "performances", "performance details", "performance evaluations", "evaluations"):
                    perf_sheet_name = name
                    break

            if not perf_sheet_name:
                if len(wb.sheetnames) == 1:
                    perf_sheet_name = wb.sheetnames[0]
                else:
                    raise HTTPException(
                        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                        detail="Excel workbook must contain a 'Performance' or 'Performance Details' sheet."
                    )

            sheet = wb[perf_sheet_name]
            first_row = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
            col_indices = {}
            for idx, val in enumerate(first_row):
                if val is not None:
                    norm = normalize_header(val)
                    mapped_field = PERFORMANCE_HEADER_MAP.get(norm)
                    if mapped_field:
                        col_indices[mapped_field] = idx + 1

            raw_rows = []
            for r in range(2, sheet.max_row + 1):
                is_blank = True
                for c in range(1, sheet.max_column + 1):
                    val = sheet.cell(row=r, column=c).value
                    if val is not None and str(val).strip() != "":
                        is_blank = False
                        break
                if is_blank:
                    continue

                row_dict = {"excel_row": r}
                original_engineer_name = None
                for idx, val in enumerate(first_row):
                    if val is not None:
                        norm = normalize_header(val)
                        if norm in ("engineername", "name"):
                            original_engineer_name = clean_val(sheet.cell(row=r, column=idx + 1).value)
                            break
                row_dict["original_engineer_name"] = original_engineer_name

                for field, col_idx in col_indices.items():
                    row_dict[field] = clean_val(sheet.cell(row=r, column=col_idx).value)

                raw_rows.append(row_dict)

            if not raw_rows:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="The Performance sheet is empty or contains no rows."
                )

            raw_sch_ids = set()
            for row in raw_rows:
                v = row.get("schedule_id")
                if v:
                    try:
                        raw_sch_ids.add(uuid_pkg.UUID(str(v).strip()))
                    except ValueError:
                        pass

            db_schedules_map = {}
            if raw_sch_ids:
                sches = db.scalars(
                    select(Schedule).where(Schedule.schedule_id.in_(list(raw_sch_ids)))
                ).all()
                db_schedules_map = {sch.schedule_id: sch for sch in sches}

            sch_engineer_ids = {sch.engineer_id for sch in db_schedules_map.values() if sch.engineer_id}
            db_engineers_map = {}
            if sch_engineer_ids:
                engs = db.scalars(
                    select(Engineer).where(Engineer.engineer_id.in_(list(sch_engineer_ids)))
                ).all()
                db_engineers_map = {eng.engineer_id: eng for eng in engs}

            errors_list = []
            duplicates_list = []
            existing_list = []
            unchanged_list = []
            valid_rows_to_insert = []
            seen_keys = set()

            total_rows = len(raw_rows)

            for row_dict in raw_rows:
                row_errors = []
                raw_pk = row_dict.get("performance_id")
                parsed_pk = None

                if raw_pk is not None and str(raw_pk).strip() != "":
                    try:
                        parsed_pk = parse_uuid_safe(raw_pk)
                    except ValueError as ve:
                        row_errors.append({"field": "Performance ID", "value": str(raw_pk), "error": str(ve)})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                actual_start_date = None
                start_val = row_dict.get("actual_start_date")
                if start_val is not None and str(start_val).strip() != "":
                    try:
                        actual_start_date = parse_date(start_val)
                    except ValueError:
                        row_errors.append({"field": "Actual Start Date", "value": str(start_val), "error": "Invalid actual start date format."})

                actual_end_date = None
                end_val = row_dict.get("actual_end_date")
                if end_val is not None and str(end_val).strip() != "":
                    try:
                        actual_end_date = parse_date(end_val)
                    except ValueError:
                        row_errors.append({"field": "Actual End Date", "value": str(end_val), "error": "Invalid actual end date format."})

                if actual_start_date and actual_end_date and actual_end_date < actual_start_date:
                    row_errors.append({"field": "Actual End Date", "value": str(end_val), "error": "actual_end_date should not be earlier than actual_start_date"})

                score = None
                score_val = row_dict.get("score")
                if score_val is not None and str(score_val).strip() != "":
                    try:
                        score = float(score_val)
                        if score < 1.0 or score > 5.0:
                            row_errors.append({"field": "Score", "value": str(score_val), "error": "Performance rating score must be between 1.0 and 5.0"})
                    except ValueError:
                        row_errors.append({"field": "Score", "value": str(score_val), "error": "Score must be a valid number."})

                escalation_val = row_dict.get("escalation")
                escalation = False
                if escalation_val is not None:
                    if isinstance(escalation_val, bool):
                        escalation = escalation_val
                    else:
                        str_esc = str(escalation_val).strip().lower()
                        if str_esc in ("true", "yes", "y", "1"):
                            escalation = True
                        elif str_esc in ("false", "no", "n", "0"):
                            escalation = False
                        else:
                            row_errors.append({"field": "Escalation", "value": str(escalation_val), "error": "Escalation must be true or false."})

                escalation_reason = row_dict.get("escalation_reason")
                if escalation and not (escalation_reason and str(escalation_reason).strip()):
                    row_errors.append({"field": "Escalation Reason", "value": "", "error": "Escalation reason is required when escalation is enabled."})

                resolved_owner_id = resolve_owner(row_dict.get("owner"))
                row_dict["owner_id"] = resolved_owner_id
                row_dict["actual_start_date"] = actual_start_date
                row_dict["actual_end_date"] = actual_end_date
                row_dict["score"] = score
                row_dict["escalation"] = escalation
                row_dict["escalation_reason"] = escalation_reason
                row_dict["feedback"] = row_dict.get("feedback")
                row_dict["attachment"] = row_dict.get("attachment")

                if row_errors:
                    row_dict["errors"] = row_errors
                    errors_list.append(row_dict)
                    continue

                # PRIMARY KEY UPDATE PATH
                if parsed_pk is not None:
                    db_pf = db.get(Performance, parsed_pk)
                    if not db_pf:
                        row_errors.append({"field": "Performance ID", "value": str(parsed_pk), "error": f"Record with Performance ID '{parsed_pk}' was not found. Cannot update non-existent record ID."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    sch = db.get(Schedule, db_pf.schedule_id)
                    eng = db.get(Engineer, sch.engineer_id) if sch else None
                    if not eng or eng.company_id != target_company_id:
                        row_errors.append({"field": "Performance ID", "value": str(parsed_pk), "error": f"Record with Performance ID '{parsed_pk}' belongs to another company."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    row_dict["resolved_engineer_name"] = eng.engineer_name
                    row_dict["orbit_id"] = eng.orbit_id

                    changes = []
                    field_specs = [
                        ("actual_start_date", "Actual Start Date", db_pf.actual_start_date),
                        ("actual_end_date", "Actual End Date", db_pf.actual_end_date),
                        ("score", "Score", db_pf.score),
                        ("escalation", "Escalation", db_pf.escalation),
                        ("escalation_reason", "Escalation Reason", db_pf.escalation_reason),
                        ("feedback", "Feedback", db_pf.feedback),
                        ("attachment", "Attachment", db_pf.attachment),
                        ("owner_id", "Owner ID", db_pf.owner_id)
                    ]

                    for f_key, f_label, cur_val in field_specs:
                        if f_key in col_indices or (f_key == "owner_id" and "owner" in col_indices):
                            new_val = row_dict.get(f_key)
                            if not values_are_equal(cur_val, new_val):
                                changes.append(f"{f_label}: '{cur_val}' -> '{new_val}'")
                                setattr(db_pf, f_key, new_val)

                    if changes:
                        db_pf.updated_at = datetime.utcnow()
                        row_dict["update_status"] = "UPDATED"
                        row_dict["changed_fields"] = "; ".join(changes)
                        existing_list.append(row_dict)
                    else:
                        row_dict["update_status"] = "UNCHANGED"
                        row_dict["changed_fields"] = "No fields modified"
                        unchanged_list.append(row_dict)

                # NEW RECORD INSERTION PATH
                else:
                    raw_sch_str = str(row_dict.get("schedule_id") or "").strip()
                    if not raw_sch_str:
                        row_errors.append({"field": "Schedule ID", "value": "", "error": "Schedule ID is required for new performance evaluation."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    parsed_sch_id = None
                    try:
                        parsed_sch_id = uuid_pkg.UUID(raw_sch_str)
                    except ValueError:
                        row_errors.append({"field": "Schedule ID", "value": raw_sch_str, "error": f"Schedule '{raw_sch_str}' was not found. Performance record was not created."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    target_sch = db_schedules_map.get(parsed_sch_id)
                    if not target_sch:
                        row_errors.append({"field": "Schedule ID", "value": raw_sch_str, "error": f"Schedule '{raw_sch_str}' was not found. Performance record was not created."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    eng = db_engineers_map.get(target_sch.engineer_id) if target_sch.engineer_id else None
                    if not eng or eng.company_id != target_company_id:
                        row_errors.append({"field": "Schedule ID", "value": raw_sch_str, "error": f"Schedule '{raw_sch_str}' has no valid engineer in the selected company."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    uploaded_orbit_id = str(row_dict.get("orbit_id") or "").strip()
                    if uploaded_orbit_id and uploaded_orbit_id != eng.orbit_id:
                        row_errors.append({"field": "Orbit ID", "value": uploaded_orbit_id, "error": f"Orbit ID {uploaded_orbit_id} does not match Schedule {parsed_sch_id}, which belongs to Orbit ID {eng.orbit_id}."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    if not actual_start_date:
                        actual_start_date = target_sch.start_date
                        row_dict["actual_start_date"] = actual_start_date

                    if score is None:
                        row_errors.append({"field": "Score", "value": "", "error": "Score is required."})

                    if row_errors:
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    row_dict["schedule_id"] = target_sch.schedule_id
                    row_dict["engineer_id"] = eng.engineer_id
                    row_dict["resolved_engineer_name"] = eng.engineer_name
                    row_dict["orbit_id"] = eng.orbit_id

                    row_key = (norm_uuid(target_sch.schedule_id), norm_date(actual_start_date))
                    if row_key in seen_keys:
                        row_dict["duplicate_key"] = f"ScheduleID: {target_sch.schedule_id}, ActualStartDate: {actual_start_date}"
                        duplicates_list.append(row_dict)
                        continue
                    seen_keys.add(row_key)

                    valid_rows_to_insert.append(row_dict)

            # Persist Inserts & Updates
            imported_count = 0
            failed_count = 0
            try:
                perfs_to_add = []
                for item in valid_rows_to_insert:
                    db_pf = Performance(
                        performance_id=uuid_pkg.uuid4(),
                        schedule_id=item["schedule_id"],
                        owner_id=item.get("owner_id"),
                        actual_start_date=item["actual_start_date"],
                        actual_end_date=item.get("actual_end_date"),
                        escalation=item.get("escalation"),
                        escalation_reason=item.get("escalation_reason"),
                        feedback=item.get("feedback"),
                        score=item.get("score"),
                        attachment=item.get("attachment"),
                        created_at=datetime.utcnow(),
                        updated_at=datetime.utcnow()
                    )
                    perfs_to_add.append(db_pf)
                if perfs_to_add:
                    db.add_all(perfs_to_add)

                db.commit()
                imported_count = len(valid_rows_to_insert) + len(existing_list)
            except Exception as insert_err:
                db.rollback()
                failed_count = len(valid_rows_to_insert) + len(existing_list)
                bulk_upload_service.update_bulk_upload(
                    db,
                    upload_id=upload_id,
                    status="FAILED",
                    failed_rows=failed_count
                )
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail=f"Database ingestion failed: {str(insert_err)}"
                )

            # Generate Report
            report_wb = openpyxl.Workbook()
            ws_summary = report_wb.active
            ws_summary.title = "Summary"
            ws_summary.append(["ORMP Performance Bulk Ingestion Report"])
            ws_summary.append([])
            ws_summary.append(["File Name", file.filename])
            ws_summary.append(["Uploaded By", current_user.full_name])
            ws_summary.append(["Upload Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            ws_summary.append(["Target Company", company.company_name])
            ws_summary.append(["Company UUID", str(company.company_id)])
            ws_summary.append([])
            ws_summary.append(["Metric", "Count"])
            ws_summary.append(["Total Rows", total_rows])
            ws_summary.append(["Inserted Records", len(valid_rows_to_insert)])
            ws_summary.append(["Updated Records", len(existing_list)])
            ws_summary.append(["Unchanged Records", len(unchanged_list)])
            ws_summary.append(["Error Rows", len(errors_list)])
            ws_summary.append(["Duplicate Rows", len(duplicates_list)])
            ws_summary.append(["Warning Rows", 0])

            for col in ws_summary.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws_summary.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)

            headers_valid = [
                "Excel Row", "Performance ID", "Schedule ID", "Orbit ID", "Engineer Name", "Score", "Actual Start", 
                "Actual End", "Escalation", "Escalation Reason", "Feedback", "Status"
            ]

            ws_valid = report_wb.create_sheet(title="Valid Records")
            ws_valid.append(headers_valid)
            for r in valid_rows_to_insert:
                ws_valid.append([
                    r["excel_row"],
                    str(r.get("performance_id") or ""),
                    str(r.get("schedule_id") or ""),
                    r.get("orbit_id"),
                    r.get("resolved_engineer_name"),
                    r.get("score"),
                    str(r.get("actual_start_date")) if r.get("actual_start_date") else "",
                    str(r.get("actual_end_date")) if r.get("actual_end_date") else "",
                    "Yes" if r.get("escalation") else "No",
                    r.get("escalation_reason"),
                    r.get("feedback"),
                    "INSERTED"
                ])

            ws_updated = report_wb.create_sheet(title="Updated Records")
            ws_updated.append(["Excel Row", "Performance ID", "Schedule ID", "Orbit ID", "Engineer Name", "Action Status", "Changed Columns", "Score", "Escalation", "Escalation Reason", "Feedback"])
            for r in existing_list:
                ws_updated.append([
                    r["excel_row"],
                    str(r.get("performance_id") or ""),
                    str(r.get("schedule_id") or ""),
                    r.get("orbit_id") or "",
                    r.get("resolved_engineer_name") or r.get("original_engineer_name") or "",
                    "UPDATED",
                    r.get("changed_fields") or "",
                    r.get("score"),
                    "Yes" if r.get("escalation") else "No",
                    r.get("escalation_reason"),
                    r.get("feedback")
                ])

            ws_unchanged = report_wb.create_sheet(title="Unchanged Records")
            ws_unchanged.append(["Excel Row", "Performance ID", "Schedule ID", "Orbit ID", "Engineer Name", "Action Status", "Details", "Score", "Escalation", "Escalation Reason", "Feedback"])
            for r in unchanged_list:
                ws_unchanged.append([
                    r["excel_row"],
                    str(r.get("performance_id") or ""),
                    str(r.get("schedule_id") or ""),
                    r.get("orbit_id") or "",
                    r.get("resolved_engineer_name") or r.get("original_engineer_name") or "",
                    "UNCHANGED",
                    "All supplied values match database",
                    r.get("score"),
                    "Yes" if r.get("escalation") else "No",
                    r.get("escalation_reason"),
                    r.get("feedback")
                ])

            ws_errors = report_wb.create_sheet(title="Errors")
            ws_errors.append(["Excel Row", "Orbit ID", "Field", "Value", "Error"])
            for r in errors_list:
                for err in r.get("errors", []):
                    ws_errors.append([
                        r["excel_row"],
                        r.get("orbit_id") or "",
                        err.get("field") or "",
                        err.get("value") or "",
                        err.get("error") or ""
                    ])

            ws_dups = report_wb.create_sheet(title="Duplicates")
            ws_dups.append(["Excel Row", "Orbit ID", "Duplicate Key", "Reason"])
            for r in duplicates_list:
                ws_dups.append([
                    r["excel_row"],
                    r.get("orbit_id") or "",
                    r.get("duplicate_key") or "",
                    "Duplicate Performance row within Excel sheet"
                ])

            ws_warn = report_wb.create_sheet(title="Warnings")
            ws_warn.append(["Excel Row", "Schedule ID", "Orbit ID", "Field", "Value", "Warning"])

            for sheet_obj in (ws_valid, ws_updated, ws_unchanged, ws_errors, ws_dups, ws_warn):
                for col in sheet_obj.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    sheet_obj.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)

            os.makedirs(TEMP_REPORTS_DIR, exist_ok=True)
            report_filename = f"validation_report_{uuid_pkg.uuid4()}.xlsx"
            report_path = os.path.join(TEMP_REPORTS_DIR, report_filename)
            report_wb.save(report_path)

            final_status = "COMPLETED"
            if len(errors_list) > 0 or len(duplicates_list) > 0:
                final_status = "COMPLETED_WITH_ERRORS"

            bulk_upload_service.update_bulk_upload(
                db,
                upload_id=upload_id,
                status=final_status,
                report_file=report_filename,
                imported_rows=imported_count,
                failed_rows=failed_count
            )

            ingested_msg = f"Ingested {len(valid_rows_to_insert)} new performance records. Updated {len(existing_list)} existing performance records. {len(unchanged_list)} records were unchanged."
            if errors_list or duplicates_list:
                ingested_msg += " Some rows had validation errors or duplicates. See the validation report for details."

            return {
                "success": True,
                "rowsProcessed": total_rows,
                "errorsCount": len(errors_list),
                "inserted": len(valid_rows_to_insert),
                "updated": len(existing_list),
                "unchanged": len(unchanged_list),
                "message": ingested_msg,
                "report_url": f"/api/upload/download-report/{report_filename}"
            }

        # =========================================================================
        # MODULE 6: UP-LEAVE
        # =========================================================================
        if module_id == "up-leave":
            start_time = time.perf_counter()
            try:
                contents = await file.read()
                wb = openpyxl.load_workbook(io.BytesIO(contents))
            except Exception:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Failed to parse Excel file. Please ensure it is a valid .xlsx file."
                )

            leave_sheet_name = None
            for name in wb.sheetnames:
                norm_name = name.strip().lower()
                if norm_name in ("leave", "leaves", "leave details", "leave records", "absences"):
                    leave_sheet_name = name
                    break

            if not leave_sheet_name:
                if len(wb.sheetnames) == 1:
                    leave_sheet_name = wb.sheetnames[0]
                else:
                    raise HTTPException(
                        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                        detail="Excel workbook must contain a 'Leave' or 'Leaves' sheet."
                    )

            sheet = wb[leave_sheet_name]
            first_row = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
            col_indices = {}
            for idx, val in enumerate(first_row):
                if val is not None:
                    norm = normalize_header(val)
                    mapped_field = LEAVE_HEADER_MAP.get(norm)
                    if mapped_field:
                        col_indices[mapped_field] = idx + 1

            raw_rows = []
            for r in range(2, sheet.max_row + 1):
                is_blank = True
                for c in range(1, sheet.max_column + 1):
                    val = sheet.cell(row=r, column=c).value
                    if val is not None and str(val).strip() != "":
                        is_blank = False
                        break
                if is_blank:
                    continue

                row_dict = {"excel_row": r}
                original_engineer_name = None
                for idx, val in enumerate(first_row):
                    if val is not None:
                        norm = normalize_header(val)
                        if norm in ("engineername", "name"):
                            original_engineer_name = clean_val(sheet.cell(row=r, column=idx + 1).value)
                            break
                row_dict["original_engineer_name"] = original_engineer_name

                for field, col_idx in col_indices.items():
                    row_dict[field] = clean_val(sheet.cell(row=r, column=col_idx).value)

                raw_rows.append(row_dict)

            if not raw_rows:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="The Leave sheet is empty or contains no rows."
                )

            unique_orbit_ids = {
                norm_str(row.get("orbit_id"))
                for row in raw_rows
                if row.get("orbit_id") and norm_str(row.get("orbit_id")) != ""
            }

            db_engineers = []
            if unique_orbit_ids:
                db_engineers = db.scalars(
                    select(Engineer).where(
                        func.lower(Engineer.orbit_id).in_(list(unique_orbit_ids)),
                        Engineer.company_id == target_company_id
                    )
                ).all()

            orbit_to_engineer = {
                norm_str(eng.orbit_id): (eng.engineer_id, eng.engineer_name)
                for eng in db_engineers
            }

            errors_list = []
            duplicates_list = []
            existing_list = []
            unchanged_list = []
            valid_rows_to_insert = []
            seen_keys = set()

            total_rows = len(raw_rows)

            for row_dict in raw_rows:
                row_errors = []
                raw_pk = row_dict.get("leave_id")
                parsed_pk = None

                if raw_pk is not None and str(raw_pk).strip() != "":
                    try:
                        parsed_pk = parse_uuid_safe(raw_pk)
                    except ValueError as ve:
                        row_errors.append({"field": "Leave ID", "value": str(raw_pk), "error": str(ve)})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                requested_date = None
                req_date_val = row_dict.get("requested_date")
                if req_date_val is not None:
                    try:
                        requested_date = parse_date(req_date_val)
                    except ValueError:
                        row_errors.append({"field": "Requested Date", "value": str(req_date_val), "error": "Invalid requested date format."})

                requested_on = None
                req_on_val = row_dict.get("requested_on")
                if req_on_val is not None:
                    try:
                        requested_on = parse_date(req_on_val)
                    except ValueError:
                        row_errors.append({"field": "Requested On", "value": str(req_on_val), "error": "Invalid requested on submission date format."})
                else:
                    requested_on = date.today()

                if requested_date and requested_on and requested_on > requested_date:
                    row_errors.append({"field": "Requested On", "value": str(req_on_val), "error": "requested_on date cannot be later than requested_date"})

                leave_type = row_dict.get("leave_type") or "Annual Leave"
                approval_status = row_dict.get("approval_status") or "Pending"
                norm_status = str(approval_status).strip().title()
                if norm_status not in ("Pending", "Approved", "Rejected", "Cancelled"):
                    norm_status = "Pending"
                approval_status = norm_status

                resolved_owner_id = resolve_owner(row_dict.get("owner"))
                row_dict["owner_id"] = resolved_owner_id
                row_dict["requested_date"] = requested_date
                row_dict["requested_on"] = requested_on
                row_dict["leave_type"] = leave_type
                row_dict["approval_status"] = approval_status

                if row_errors:
                    row_dict["errors"] = row_errors
                    errors_list.append(row_dict)
                    continue

                # PRIMARY KEY UPDATE PATH
                if parsed_pk is not None:
                    db_lv = db.get(Leave, parsed_pk)
                    if not db_lv:
                        row_errors.append({"field": "Leave ID", "value": str(parsed_pk), "error": f"Record with Leave ID '{parsed_pk}' was not found. Cannot update non-existent record ID."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    eng = db.get(Engineer, db_lv.engineer_id)
                    if not eng or eng.company_id != target_company_id:
                        row_errors.append({"field": "Leave ID", "value": str(parsed_pk), "error": f"Record with Leave ID '{parsed_pk}' belongs to another company."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    row_dict["resolved_engineer_name"] = eng.engineer_name
                    row_dict["orbit_id"] = eng.orbit_id

                    changes = []
                    field_specs = [
                        ("leave_type", "Leave Type", db_lv.leave_type),
                        ("requested_date", "Requested Date", db_lv.requested_date),
                        ("requested_on", "Requested On", db_lv.requested_on),
                        ("approval_status", "Approval Status", db_lv.approval_status),
                        ("owner_id", "Owner ID", db_lv.owner_id)
                    ]

                    for f_key, f_label, cur_val in field_specs:
                        if f_key in col_indices or (f_key == "owner_id" and "owner" in col_indices):
                            new_val = row_dict.get(f_key)
                            if not values_are_equal(cur_val, new_val):
                                changes.append(f"{f_label}: '{cur_val}' -> '{new_val}'")
                                setattr(db_lv, f_key, new_val)

                    if changes:
                        db_lv.updated_at = datetime.utcnow()
                        row_dict["update_status"] = "UPDATED"
                        row_dict["changed_fields"] = "; ".join(changes)
                        existing_list.append(row_dict)
                    else:
                        row_dict["update_status"] = "UNCHANGED"
                        row_dict["changed_fields"] = "No fields modified"
                        unchanged_list.append(row_dict)

                # NEW RECORD INSERTION PATH
                else:
                    orbit_id = row_dict.get("orbit_id")
                    if not orbit_id:
                        row_errors.append({"field": "Orbit ID", "value": "", "error": "Orbit ID is required for new leave record."})

                    if not requested_date:
                        row_errors.append({"field": "Requested Date", "value": "", "error": "Requested Date is required."})

                    if row_errors:
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    eng_info = orbit_to_engineer.get(norm_str(orbit_id))
                    if not eng_info:
                        row_errors.append({"field": "Orbit ID", "value": str(orbit_id), "error": f"Engineer with Orbit ID '{orbit_id}' does not exist in the selected company."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    engineer_id, resolved_engineer_name = eng_info
                    row_dict["engineer_id"] = engineer_id
                    row_dict["resolved_engineer_name"] = resolved_engineer_name

                    lt_clean = norm_str(leave_type or "Annual Leave")
                    row_key = (norm_uuid(engineer_id), norm_date(requested_date), lt_clean)

                    if row_key in seen_keys:
                        row_dict["duplicate_key"] = f"OrbitID: {orbit_id}, RequestedDate: {requested_date}, LeaveType: {leave_type}"
                        duplicates_list.append(row_dict)
                        continue
                    seen_keys.add(row_key)

                    valid_rows_to_insert.append(row_dict)

            # Persist Inserts & Updates
            imported_count = 0
            failed_count = 0
            try:
                leaves_to_add = []
                for item in valid_rows_to_insert:
                    db_lv = Leave(
                        leave_id=uuid_pkg.uuid4(),
                        engineer_id=item["engineer_id"],
                        owner_id=item.get("owner_id"),
                        leave_type=item["leave_type"],
                        requested_date=item["requested_date"],
                        requested_on=item["requested_on"],
                        approval_status=item["approval_status"],
                        created_at=datetime.utcnow(),
                        updated_at=datetime.utcnow()
                    )
                    leaves_to_add.append(db_lv)
                if leaves_to_add:
                    db.add_all(leaves_to_add)

                db.commit()
                imported_count = len(valid_rows_to_insert) + len(existing_list)
            except Exception as insert_err:
                db.rollback()
                failed_count = len(valid_rows_to_insert) + len(existing_list)
                bulk_upload_service.update_bulk_upload(
                    db,
                    upload_id=upload_id,
                    status="FAILED",
                    failed_rows=failed_count
                )
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail=f"Database ingestion failed: {str(insert_err)}"
                )

            # Generate Report
            report_wb = openpyxl.Workbook()
            ws_summary = report_wb.active
            ws_summary.title = "Summary"
            ws_summary.append(["ORMP Leave Bulk Ingestion Report"])
            ws_summary.append([])
            ws_summary.append(["File Name", file.filename])
            ws_summary.append(["Uploaded By", current_user.full_name])
            ws_summary.append(["Upload Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            ws_summary.append(["Target Company", company.company_name])
            ws_summary.append(["Company UUID", str(company.company_id)])
            ws_summary.append([])
            ws_summary.append(["Metric", "Count"])
            ws_summary.append(["Total Rows", total_rows])
            ws_summary.append(["Inserted Records", len(valid_rows_to_insert)])
            ws_summary.append(["Updated Records", len(existing_list)])
            ws_summary.append(["Unchanged Records", len(unchanged_list)])
            ws_summary.append(["Error Rows", len(errors_list)])
            ws_summary.append(["Duplicate Rows", len(duplicates_list)])
            ws_summary.append(["Warning Rows", 0])

            for col in ws_summary.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws_summary.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)

            headers_valid = [
                "Excel Row", "Leave ID", "Orbit ID", "Engineer Name", "Leave Type", 
                "Requested Date", "Requested On", "Approval Status", "Status"
            ]

            ws_valid = report_wb.create_sheet(title="Valid Records")
            ws_valid.append(headers_valid)
            for r in valid_rows_to_insert:
                ws_valid.append([
                    r["excel_row"],
                    str(r.get("leave_id") or ""),
                    r.get("orbit_id"),
                    r.get("resolved_engineer_name"),
                    r.get("leave_type"),
                    str(r.get("requested_date")) if r.get("requested_date") else "",
                    str(r.get("requested_on")) if r.get("requested_on") else "",
                    r.get("approval_status"),
                    "INSERTED"
                ])

            ws_updated = report_wb.create_sheet(title="Updated Records")
            ws_updated.append(["Excel Row", "Leave ID", "Orbit ID", "Engineer Name", "Action Status", "Changed Columns", "Leave Type", "Requested Date", "Requested On", "Approval Status"])
            for r in existing_list:
                ws_updated.append([
                    r["excel_row"],
                    str(r.get("leave_id") or ""),
                    r.get("orbit_id") or "",
                    r.get("resolved_engineer_name") or r.get("original_engineer_name") or "",
                    "UPDATED",
                    r.get("changed_fields") or "",
                    r.get("leave_type"),
                    str(r.get("requested_date")) if r.get("requested_date") else "",
                    str(r.get("requested_on")) if r.get("requested_on") else "",
                    r.get("approval_status")
                ])

            ws_unchanged = report_wb.create_sheet(title="Unchanged Records")
            ws_unchanged.append(["Excel Row", "Leave ID", "Orbit ID", "Engineer Name", "Action Status", "Details", "Leave Type", "Requested Date", "Requested On", "Approval Status"])
            for r in unchanged_list:
                ws_unchanged.append([
                    r["excel_row"],
                    str(r.get("leave_id") or ""),
                    r.get("orbit_id") or "",
                    r.get("resolved_engineer_name") or r.get("original_engineer_name") or "",
                    "UNCHANGED",
                    "All supplied values match database",
                    r.get("leave_type"),
                    str(r.get("requested_date")) if r.get("requested_date") else "",
                    str(r.get("requested_on")) if r.get("requested_on") else "",
                    r.get("approval_status")
                ])

            ws_errors = report_wb.create_sheet(title="Errors")
            ws_errors.append(["Excel Row", "Orbit ID", "Field", "Value", "Error"])
            for r in errors_list:
                for err in r.get("errors", []):
                    ws_errors.append([
                        r["excel_row"],
                        r.get("orbit_id") or "",
                        err.get("field") or "",
                        err.get("value") or "",
                        err.get("error") or ""
                    ])

            ws_dups = report_wb.create_sheet(title="Duplicates")
            ws_dups.append(["Excel Row", "Orbit ID", "Duplicate Key", "Reason"])
            for r in duplicates_list:
                ws_dups.append([
                    r["excel_row"],
                    r.get("orbit_id") or "",
                    r.get("duplicate_key") or "",
                    "Duplicate Leave row within Excel sheet"
                ])

            ws_warn = report_wb.create_sheet(title="Warnings")
            ws_warn.append(["Excel Row", "Orbit ID", "Field", "Value", "Warning"])

            for sheet_obj in (ws_valid, ws_updated, ws_unchanged, ws_errors, ws_dups, ws_warn):
                for col in sheet_obj.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    sheet_obj.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)

            os.makedirs(TEMP_REPORTS_DIR, exist_ok=True)
            report_filename = f"validation_report_{uuid_pkg.uuid4()}.xlsx"
            report_path = os.path.join(TEMP_REPORTS_DIR, report_filename)
            report_wb.save(report_path)

            final_status = "COMPLETED"
            if len(errors_list) > 0 or len(duplicates_list) > 0:
                final_status = "COMPLETED_WITH_ERRORS"

            bulk_upload_service.update_bulk_upload(
                db,
                upload_id=upload_id,
                status=final_status,
                report_file=report_filename,
                imported_rows=imported_count,
                failed_rows=failed_count
            )

            ingested_msg = f"Ingested {len(valid_rows_to_insert)} new leave records. Updated {len(existing_list)} existing leave records. {len(unchanged_list)} records were unchanged."
            if errors_list or duplicates_list:
                ingested_msg += " Some rows had validation errors or duplicates. See the validation report for details."

            return {
                "success": True,
                "rowsProcessed": total_rows,
                "errorsCount": len(errors_list),
                "inserted": len(valid_rows_to_insert),
                "updated": len(existing_list),
                "unchanged": len(unchanged_list),
                "message": ingested_msg,
                "report_url": f"/api/upload/download-report/{report_filename}"
            }

        # =========================================================================
        # MODULE 7: UP-ENGINEERS
        # =========================================================================
        if module_id == "up-engineers":
            start_time = time.perf_counter()
            try:
                contents = await file.read()
                wb = openpyxl.load_workbook(io.BytesIO(contents))
            except Exception:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Failed to parse Excel file. Please ensure it is a valid .xlsx file."
                )

            engineer_sheet_name = None
            for name in wb.sheetnames:
                if name.strip().lower() in ("engineer", "engineers", "roster"):
                    engineer_sheet_name = name
                    break

            if not engineer_sheet_name:
                if len(wb.sheetnames) == 1:
                    engineer_sheet_name = wb.sheetnames[0]
                else:
                    raise HTTPException(
                        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                        detail="Excel workbook must contain an Engineer sheet."
                    )

            sheet = wb[engineer_sheet_name]
            last_data_row = 1
            for r in range(sheet.max_row, 1, -1):
                row_has_data = False
                for c in range(1, sheet.max_column + 1):
                    val = sheet.cell(row=r, column=c).value
                    if val is not None and str(val).strip() != "":
                        row_has_data = True
                        break
                if row_has_data:
                    last_data_row = r
                    break

            if last_data_row <= 1:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="The Engineer sheet is empty or contains no rows."
                )

            first_row = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
            col_indices = {}
            for idx, val in enumerate(first_row):
                if val is not None:
                    mapped_field = map_engineer_header(val)
                    if mapped_field:
                        col_indices[mapped_field] = idx + 1

            errors_list = []
            duplicates_list = []
            existing_list = []
            unchanged_list = []
            valid_rows_to_insert = []
            seen_orbit_ids = set()

            total_rows = last_data_row - 1

            all_fields = {
                "engineer_id", "engineer_name", "goes_by", "employee_id", "orbit_id", "level",
                "date_of_joining", "primary_tool", "customer_experience",
                "industry_experience", "status", "email", "phone_number"
            }

            for r in range(2, last_data_row + 1):
                row_dict = {}
                row_dict["excel_row"] = r
                for field, col_idx in col_indices.items():
                    row_dict[field] = clean_val(sheet.cell(row=r, column=col_idx).value)

                for field in all_fields:
                    if field not in row_dict:
                        row_dict[field] = None

                if row_dict.get("employee_id") is not None:
                    emp_v = row_dict["employee_id"]
                    if isinstance(emp_v, float):
                        if emp_v.is_integer():
                            row_dict["employee_id"] = str(int(emp_v))
                        else:
                            row_dict["employee_id"] = str(emp_v).strip()
                    else:
                        row_dict["employee_id"] = str(emp_v).strip()

                row_errors = []

                email_val = row_dict.get("email")
                if email_val:
                    email_str = str(email_val).strip()
                    if not EMAIL_REGEX.match(email_str):
                        row_errors.append({"field": "Email", "value": email_str, "error": "Invalid email format"})
                    else:
                        row_dict["email"] = email_str

                phone_val = row_dict.get("phone_number")
                if phone_val:
                    phone_str = str(phone_val).strip()
                    if not re.match(r"^[+\d\s().-]{3,30}$", phone_str):
                        row_errors.append({"field": "Phone Number", "value": phone_str, "error": "Phone number is invalid or too long"})
                    else:
                        row_dict["phone_number"] = phone_str

                normalized_cust_exp = None
                if row_dict.get("customer_experience") is not None:
                    try:
                        normalized_cust_exp = parse_experience(row_dict["customer_experience"])
                    except ValueError:
                        row_errors.append({"field": "Customer Experience", "value": str(row_dict["customer_experience"]), "error": "LAM Experience must be numeric or 'X Years'"})

                normalized_ind_exp = None
                if row_dict.get("industry_experience") is not None:
                    try:
                        normalized_ind_exp = parse_experience(row_dict["industry_experience"])
                    except ValueError:
                        row_errors.append({"field": "Industry Experience", "value": str(row_dict["industry_experience"]), "error": "Industry Experience must be numeric or 'X Years'"})

                normalized_date = None
                if row_dict.get("date_of_joining") is not None:
                    try:
                        normalized_date = parse_date(row_dict["date_of_joining"])
                    except ValueError:
                        row_errors.append({"field": "Date of Joining", "value": str(row_dict["date_of_joining"]), "error": "Date of Joining is invalid"})

                row_dict["date_of_joining"] = normalized_date
                row_dict["customer_experience"] = normalized_cust_exp
                row_dict["industry_experience"] = normalized_ind_exp

                if row_errors:
                    row_dict["errors"] = row_errors
                    errors_list.append(row_dict)
                    continue

                raw_pk = row_dict.get("engineer_id")
                parsed_pk = None

                if raw_pk is not None and str(raw_pk).strip() != "":
                    try:
                        parsed_pk = parse_uuid_safe(raw_pk)
                    except ValueError as ve:
                        row_errors.append({"field": "Engineer ID", "value": str(raw_pk), "error": str(ve)})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                # PRIMARY KEY UPDATE PATH
                if parsed_pk is not None:
                    db_exist = db.get(Engineer, parsed_pk)
                    if not db_exist:
                        row_errors.append({"field": "Engineer ID", "value": str(parsed_pk), "error": f"Engineer record with ID '{parsed_pk}' was not found. Cannot update non-existent record ID."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    # Verify Tenant
                    if db_exist.company_id != target_company_id:
                        row_errors.append({"field": "Engineer ID", "value": str(parsed_pk), "error": f"Engineer record with ID '{parsed_pk}' belongs to another company."})
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    row_dict["engineer_id"] = db_exist.engineer_id
                    changes = []
                    field_specs = [
                        ("engineer_name", "Engineer Name", db_exist.engineer_name),
                        ("goes_by", "Goes By", db_exist.goes_by),
                        ("employee_id", "Employee ID", db_exist.lam_id),
                        ("orbit_id", "Orbit ID", db_exist.orbit_id),
                        ("level", "Level", db_exist.level),
                        ("date_of_joining", "Date of Joining", db_exist.date_of_joining),
                        ("primary_tool", "Primary Tool", db_exist.primary_tool_type),
                        ("customer_experience", "Customer Experience", float(db_exist.lam_experience) if db_exist.lam_experience is not None else None),
                        ("industry_experience", "Industry Experience", float(db_exist.industry_experience) if db_exist.industry_experience is not None else None),
                        ("status", "Status", db_exist.status),
                        ("email", "Email", db_exist.email),
                        ("phone_number", "Phone Number", db_exist.phone_number)
                    ]

                    for f_key, f_label, cur_val in field_specs:
                        if f_key in col_indices:
                            new_val = row_dict.get(f_key)
                            if not values_are_equal(cur_val, new_val):
                                changes.append(f"{f_label}: '{cur_val}' -> '{new_val}'")
                                if f_key == "employee_id":
                                    db_exist.lam_id = new_val
                                elif f_key == "primary_tool":
                                    db_exist.primary_tool_type = new_val
                                elif f_key == "customer_experience":
                                    db_exist.lam_experience = new_val
                                else:
                                    setattr(db_exist, f_key, new_val)

                    if changes:
                        db_exist.updated_at = datetime.utcnow()
                        row_dict["update_status"] = "UPDATED"
                        row_dict["changed_fields"] = "; ".join(changes)
                        existing_list.append(row_dict)
                    else:
                        row_dict["update_status"] = "UNCHANGED"
                        row_dict["changed_fields"] = "No fields modified"
                        unchanged_list.append(row_dict)

                # NO PRIMARY KEY -> NEW ROW ONLY (DO NOT UPDATE BY ORBIT_ID)
                else:
                    if not row_dict.get("engineer_name"):
                        row_errors.append({"field": "Engineer Name", "value": "", "error": "Engineer Name is required"})

                    if not row_dict.get("orbit_id"):
                        row_errors.append({"field": "Orbit ID", "value": "", "error": "Orbit ID is required"})

                    if row_errors:
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue

                    o_id = str(row_dict["orbit_id"]).strip()
                    if o_id in seen_orbit_ids:
                        row_dict["duplicate_key"] = f"Orbit ID: {o_id}"
                        duplicates_list.append(row_dict)
                        continue
                    seen_orbit_ids.add(o_id)

                    # Check if engineer with this Orbit ID already exists in DB
                    db_exist = db.scalars(
                        select(Engineer).where(
                            func.lower(Engineer.orbit_id) == norm_str(o_id),
                            Engineer.company_id == target_company_id
                        )
                    ).first()

                    if db_exist:
                        row_errors.append({
                            "field": "Orbit ID",
                            "value": o_id,
                            "error": f"Engineer with Orbit ID '{o_id}' already exists in database. To update an existing engineer record, supply the engineer_id."
                        })
                        row_dict["errors"] = row_errors
                        errors_list.append(row_dict)
                        continue
                    else:
                        valid_rows_to_insert.append(row_dict)

            # Persist Inserts & Updates
            imported_count = 0
            updated_count = 0
            failed_count = 0
            try:
                for item in valid_rows_to_insert:
                    db_engineer = Engineer(
                        engineer_id=uuid_pkg.uuid4(),
                        company_id=target_company_id,
                        engineer_name=item["engineer_name"],
                        goes_by=item.get("goes_by"),
                        lam_id=item.get("employee_id"),
                        orbit_id=item["orbit_id"],
                        level=item.get("level"),
                        date_of_joining=item.get("date_of_joining"),
                        primary_tool_type=item.get("primary_tool"),
                        lam_experience=item.get("customer_experience"),
                        industry_experience=item.get("industry_experience"),
                        status=item.get("status") or "Active",
                        email=item.get("email"),
                        phone_number=item.get("phone_number"),
                        created_at=datetime.utcnow(),
                        updated_at=datetime.utcnow()
                    )
                    db.add(db_engineer)

                db.commit()
                imported_count = len(valid_rows_to_insert)
                updated_count = len(existing_list)
            except Exception as insert_err:
                db.rollback()
                failed_count = len(valid_rows_to_insert) + len(existing_list)
                bulk_upload_service.update_bulk_upload(
                    db,
                    upload_id=upload_id,
                    status="FAILED",
                    failed_rows=failed_count
                )
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail=f"Database ingestion failed: {str(insert_err)}"
                )

            # Generate Report
            report_wb = openpyxl.Workbook()
            ws_summary = report_wb.active
            ws_summary.title = "Summary"
            ws_summary.append(["ORMP Bulk Ingestion Invalidation & Validation Report"])
            ws_summary.append([])
            ws_summary.append(["File Name", file.filename])
            ws_summary.append(["Uploaded By", current_user.full_name])
            ws_summary.append(["Upload Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            ws_summary.append(["Target Company", company.company_name])
            ws_summary.append(["Company UUID", str(company.company_id)])
            ws_summary.append([])
            ws_summary.append(["Metric", "Count"])
            ws_summary.append(["Total Rows", total_rows])
            ws_summary.append(["Inserted Records", len(valid_rows_to_insert)])
            ws_summary.append(["Updated Records", len(existing_list)])
            ws_summary.append(["Unchanged Records", len(unchanged_list)])
            ws_summary.append(["Error Rows", len(errors_list)])
            ws_summary.append(["Duplicate Rows", len(duplicates_list)])
            ws_summary.append(["Warning Rows", 0])

            for col in ws_summary.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws_summary.column_dimensions[col[0].column_letter].width = max(max_len + 3, 14)

            headers = [
                "engineer_id", "engineer_name", "goes_by", "employee_id", "orbit_id", "level", 
                "date_of_joining", "primary_tool", "customer_experience", "industry_experience", 
                "status", "email", "phone_number"
            ]

            def add_sheet_data(ws, rows, include_errors=False):
                row_headers = headers.copy()
                if include_errors:
                    row_headers.append("errors")
                ws.append(row_headers)
                for r in rows:
                    row_vals = [
                        str(r.get("engineer_id") or ""),
                        r.get("engineer_name"),
                        r.get("goes_by"),
                        r.get("employee_id"),
                        r.get("orbit_id"),
                        r.get("level"),
                        str(r.get("date_of_joining")) if r.get("date_of_joining") else "",
                        r.get("primary_tool"),
                        r.get("customer_experience"),
                        r.get("industry_experience"),
                        r.get("status"),
                        r.get("email"),
                        r.get("phone_number")
                    ]
                    if include_errors:
                        err_msgs = [e["error"] if isinstance(e, dict) else str(e) for e in r.get("errors", [])]
                        row_vals.append(", ".join(err_msgs))
                    ws.append(row_vals)
                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)

            # 1. Valid Records
            ws_valid = report_wb.create_sheet(title="Valid Records")
            add_sheet_data(ws_valid, valid_rows_to_insert)

            # 2. Updated Records
            ws_exist = report_wb.create_sheet(title="Updated Records")
            ws_exist.append([
                "Excel Row", "Engineer ID", "Orbit ID", "Engineer Name", "Action Status", "Changed Diffs",
                "Goes By", "Employee ID", "Level", "Date of Joining", "Primary Tool",
                "Customer Exp", "Industry Exp", "Status", "Email", "Phone Number"
            ])
            for r in existing_list:
                ws_exist.append([
                    r.get("excel_row"),
                    str(r.get("engineer_id") or ""),
                    r.get("orbit_id"),
                    r.get("engineer_name"),
                    r.get("update_status", "UPDATED"),
                    r.get("changed_fields", "Modified values updated in database"),
                    r.get("goes_by"),
                    r.get("employee_id"),
                    r.get("level"),
                    str(r.get("date_of_joining")) if r.get("date_of_joining") else "",
                    r.get("primary_tool"),
                    r.get("customer_experience"),
                    r.get("industry_experience"),
                    r.get("status"),
                    r.get("email"),
                    r.get("phone_number")
                ])
            for col in ws_exist.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws_exist.column_dimensions[col[0].column_letter].width = max(max_len + 3, 14)

            # 3. Unchanged Records
            ws_un = report_wb.create_sheet(title="Unchanged Records")
            ws_un.append(["Excel Row", "Engineer ID", "Orbit ID", "Engineer Name", "Action Status", "Details"])
            for r in unchanged_list:
                ws_un.append([
                    r.get("excel_row"),
                    str(r.get("engineer_id") or ""),
                    r.get("orbit_id"),
                    r.get("engineer_name"),
                    "UNCHANGED",
                    "All supplied engineer fields match database record"
                ])
            for col in ws_un.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws_un.column_dimensions[col[0].column_letter].width = max(max_len + 3, 14)

            # 4. Errors
            ws_err = report_wb.create_sheet(title="Errors")
            ws_err.append(["Excel Row", "Orbit ID", "Field", "Value", "Error"])
            for r in errors_list:
                for err in r.get("errors", []):
                    if isinstance(err, dict):
                        ws_err.append([
                            r["excel_row"],
                            r.get("orbit_id") or "",
                            err.get("field") or "",
                            err.get("value") or "",
                            err.get("error") or ""
                        ])
                    else:
                        ws_err.append([
                            r["excel_row"],
                            r.get("orbit_id") or "",
                            "",
                            "",
                            str(err)
                        ])

            # 5. Duplicates
            ws_dups = report_wb.create_sheet(title="Duplicates")
            add_sheet_data(ws_dups, duplicates_list)

            os.makedirs(TEMP_REPORTS_DIR, exist_ok=True)
            report_filename = f"validation_report_{uuid_pkg.uuid4()}.xlsx"
            report_path = os.path.join(TEMP_REPORTS_DIR, report_filename)
            report_wb.save(report_path)

            final_status = "COMPLETED"
            if len(errors_list) > 0 or len(duplicates_list) > 0:
                final_status = "COMPLETED_WITH_ERRORS"

            bulk_upload_service.update_bulk_upload(
                db,
                upload_id=upload_id,
                status=final_status,
                report_file=report_filename,
                imported_rows=imported_count + updated_count,
                failed_rows=failed_count
            )

            msg_parts = []
            if len(valid_rows_to_insert) > 0:
                msg_parts.append(f"Ingested {len(valid_rows_to_insert)} new engineer records.")
            if len(existing_list) > 0:
                msg_parts.append(f"Updated {len(existing_list)} existing engineer records with latest uploaded data.")
            if len(unchanged_list) > 0:
                msg_parts.append(f"Preserved {len(unchanged_list)} unchanged engineer records.")
            if not msg_parts:
                ingested_msg = "No engineer records processed."
            else:
                ingested_msg = " ".join(msg_parts)

            if errors_list or duplicates_list:
                ingested_msg += " Some rows had validation errors or duplicates. See the validation report for details."

            return {
                "success": True,
                "rowsProcessed": total_rows,
                "errorsCount": len(errors_list),
                "inserted": len(valid_rows_to_insert),
                "updated": len(existing_list),
                "unchanged": len(unchanged_list),
                "message": ingested_msg,
                "report_url": f"/api/upload/download-report/{report_filename}"
            }

    except HTTPException as he:
        db.rollback()
        bulk_upload_service.update_bulk_upload(
            db,
            upload_id=upload_id,
            status="FAILED"
        )
        raise he
    except Exception as e:
        db.rollback()
        logger.error("Error during bulk upload: %s", str(e), exc_info=True)
        bulk_upload_service.update_bulk_upload(
            db,
            upload_id=upload_id,
            status="FAILED"
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Internal server error occurred during bulk upload."
        )

@router.get("/download-report/{report_name}")
def download_report(
    report_name: str,
    current_user: User = Depends(get_current_user)
):
    """
    Download a validation report Excel file.
    """
    # Prevent path traversal
    clean_name = os.path.basename(report_name)
    report_path = os.path.join(TEMP_REPORTS_DIR, clean_name)
    
    if not os.path.exists(report_path):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Validation report not found or expired."
        )
    
    return FileResponse(
        path=report_path,
        filename=f"ORMP_Validation_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

from app.schemas.pagination import PaginatedResponse

@router.get("/history", response_model=PaginatedResponse[BulkUploadResponse])
def get_upload_history(
    company_id: Optional[UUID] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # Enforce company isolation / Global Admin tenant filtering
    company_id = enforce_company_isolation(db, current_user, company_id)
    res = bulk_upload_service.get_bulk_uploads_history_paginated(
        db,
        company_id=company_id if isinstance(company_id, UUID) else None,
        page=page,
        page_size=page_size
    )
    return PaginatedResponse[BulkUploadResponse](
        items=[bulk_upload_service.map_to_response(db, u) for u in res["items"]],
        page=res["page"],
        page_size=res["page_size"],
        total=res["total"],
        total_pages=res["total_pages"]
    )

@router.get("/history/{upload_id}", response_model=BulkUploadResponse)
def get_upload_detail(
    upload_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    upload = bulk_upload_service.get_bulk_upload_by_id(db, upload_id)
    if not upload:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Upload record not found."
        )
    # Enforce IDOR protection: user can only access history records matching their tenant (or any if Global Admin)
    enforce_company_isolation(db, current_user, upload.company_id)
    return bulk_upload_service.map_to_response(db, upload)
